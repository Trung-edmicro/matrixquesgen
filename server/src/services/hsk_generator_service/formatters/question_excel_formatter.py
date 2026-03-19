# hsk_core/formatters/question_excel_formatter.py
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Alignment
import random

# --- CÁC HÀM ĐIỀN DỮ LIỆU VÀO EXCEL ---
# Sheet TN PA đúng (img) (HSK1)
def populate_individual_image_matching(worksheet, data: list):
    """
    Điền dữ liệu cho sheet 'TN PA đúng (img) (HSK1)'.
    Key JSON: individual_image_matching
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    start_row = 2  # Bắt đầu điền từ hàng thứ 2
    for i, question in enumerate(data, start=start_row):
        # Cột F: Mô tả ảnh + script
        image_options_str = "\n".join(question.get('image_options', []))
        script_chinese = question.get('script_chinese', '')
        content_f = f"{image_options_str}\n\n{script_chinese}"
        
        worksheet[f'D{i}'] = question.get('do_kho', 'NB')
        worksheet[f'F{i}'] = content_f
        # Cột G: Số lượng lựa chọn (cố định là 4)
        worksheet[f'G{i}'] = 4 
        
        # Cột H: Đáp án
        worksheet[f'H{i}'] = question.get('correct_answer')
    print(f"   ✅ Hoàn thành điền {len(data)} câu hỏi.")

# Sheet ĐS (img) HSK1 và ĐS Ko phụ đề (img) HSK1
def populate_true_false_from_array(worksheet, data: list):
    """
    Điền dữ liệu cho sheet Đúng/Sai từ một mảng JSON.
    Hàm này được dùng chung cho cả 'ĐS (img) HSK1' và 'ĐS Ko phụ đề (img) HSK1'.
    """
    print(f"   -> Đang điền dữ liệu Đúng/Sai vào sheet: {worksheet.title}")
    
    start_row = 2
    for i, item in enumerate(data, start=start_row):
        kind = item.get("kind")
        
        # Xử lý chuỗi hiển thị và lời giải dựa trên 'kind'
        if kind == "script_hinh_anh":
            content_f = f"Ảnh: {item.get('image_des', '')}\nScript: {item.get('script', '')}"
            explanation = f"{item.get('explanation', '')}\n\nPhụ đề:\n{item.get('script', '')}\n{item.get('pinyin', '')}\n\nTạm dịch: {item.get('translation', '')}"
        else:  # tuvung_hinh_anh
            content_f = f"Ảnh: {item.get('image_des', '')}\nTừ vựng: {item.get('script', '')}\nPinyin: {item.get('pinyin', '')}"
            explanation = f"{item.get('explanation', '')}\n\nTạm dịch: {item.get('translation', '')}"
            
        # Lấy độ khó từ item, mặc định là NB nếu không có
        do_kho = item.get("do_kho", "NB")

        # Điền vào Excel
        worksheet.cell(row=i, column=3).value = "DS"
        worksheet.cell(row=i, column=4).value = do_kho
        worksheet.cell(row=i, column=6).value = content_f
        worksheet.cell(row=i, column=8).value = f"đúng sai: {item.get('correct_answer')}"
        worksheet.cell(row=i, column=9).value = explanation
        
    print(f"   ✅ Hoàn thành điền {len(data)} câu hỏi.")

# Sheet TN PA đúng (img) (HL) (HSK1)
def populate_shared_image_comprehension(worksheet, data: dict):
    """
    Điền dữ liệu cho sheet 'TN PA đúng (img) (HL) (HSK1)'.
    Key JSON: shared_image_comprehension
    Sửa đổi: Merge ô cột B cho học liệu, điền script vào cột E.
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    
    questions = data.get('questions', [])
    if not questions:
        print("   ⚠️ Cảnh báo: Không có câu hỏi nào để điền cho dạng này.")
        return

    # --- Cấu hình hàng bắt đầu và số lượng câu hỏi ---
    start_row = 2  # Giả sử tiêu đề ở hàng 1, dữ liệu bắt đầu từ hàng 2
    num_questions = len(questions)
    end_row = start_row + num_questions - 1

    # --- 1. Xử lý Học liệu chung (Cột B, merged) ---
    shared_material = data.get('shared_material', [])
    shared_material_str = "\n".join(
        f"{desc}" for idx, desc in enumerate(shared_material)
    )
    
    # Merge các ô từ start_row đến end_row trong cột B
    # Ví dụ: nếu có 4 câu, sẽ merge từ B2 đến B5
    if num_questions > 0:
        merge_range = f'B{start_row}:B{end_row}'
        worksheet.merge_cells(merge_range)
        
        # Gán giá trị vào ô trên cùng bên trái của vùng đã merge
        top_left_cell = worksheet[f'B{start_row}']
        top_left_cell.value = shared_material_str
        
        # Căn chỉnh nội dung lên trên và tự động xuống dòng cho đẹp
        top_left_cell.alignment = Alignment(vertical='top', wrap_text=True)

    # --- 2. Điền script (Cột E) và đáp án (Cột H) ---
    for i, question in enumerate(questions):
        current_row = start_row + i
        
        # Cột D: Độ khó
        worksheet[f'D{current_row}'] = question.get('do_kho', 'NB')
        
        # Cột E: Script câu hỏi
        worksheet[f'E{current_row}'] = question.get('script_chinese', '')

        # Cột G: Số lượng lựa chọn (cố định là 4)
        worksheet[f'G{current_row}'] = 4
        
        # Cột H: Đáp án
        worksheet[f'H{current_row}'] = question.get('correct_answer')
        
    print(f"   ✅ Hoàn thành điền học liệu và {num_questions} câu hỏi.")

# Sheet TN PA đúng (HSK1)
def populate_reading_comprehension_choice(worksheet, data: list):
    """
    Điền dữ liệu cho sheet 'TN PA đúng (HSK1)'.
    Key JSON: reading_comprehension_choice
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    start_row = 2
    for i, question in enumerate(data, start=start_row):
        # Cột F: Ngữ cảnh + Câu hỏi
        context = question.get('context_chinese', '')
        query = question.get('query_chinese', '')
        worksheet[f'D{i}'] = question.get('do_kho', 'NB')
        worksheet[f'F{i}'] = f"{context}\n{query}"
        
        # Cột G: Các lựa chọn đáp án
        options = question.get('answer_options', [])
        options_str = "\n".join(
            f"{opt.get('chinese_text', '')}<br/>{opt.get('pinyin', '')}" for opt in options
        )
        worksheet[f'G{i}'] = options_str

        # Cột H: Đáp án đúng
        worksheet[f'H{i}'] = question.get('correct_answer')
    print(f"   ✅ Hoàn thành điền {len(data)} câu hỏi.")

# Sheet TN chọn ảnh (img) (HL) (HSK1)
def populate_image_matching_shared(worksheet, data: dict):
    """
    Điền dữ liệu cho sheet 'TN chọn ảnh đúng (HL) (HSK1)'.
    Key JSON: image_matching_questions
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    questions = data.get('questions', [])
    if not questions:
        print("   ⚠️ Cảnh báo: Không có câu hỏi nào để điền.")
        return

    start_row = 2
    num_questions = len(questions)
    end_row = start_row + num_questions - 1

    # Cột B: Học liệu chung (merge)
    shared_material = data.get('shared_material', [])
    shared_material_str = "\n".join(
        f"{desc}" for idx, desc in enumerate(shared_material)
    )
    if num_questions > 0:
        worksheet.merge_cells(f'B{start_row}:B{end_row}')
        top_cell = worksheet[f'B{start_row}']
        top_cell.value = shared_material_str
        top_cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Cột F và H: Câu hỏi và đáp án
    for i, question in enumerate(questions):
        current_row = start_row + i
        # Cột F: Câu hỏi
        chinese = question.get('question_text_chinese', '')
        pinyin = question.get('pinyin', '')
        worksheet[f'F{current_row}'] = f"{chinese} （ ）\n{pinyin}"
        # Cột H: Đáp án
        worksheet[f'H{current_row}'] = question.get('correct_answer')
        # Cột D: Độ khó
        worksheet[f'D{current_row}'] = question.get('do_kho', 'NB')

        # Cột G: Số lượng lựa chọn (cố định là 4)
        worksheet[f'G{current_row}'] = 4
    print(f"   ✅ Hoàn thành điền học liệu và {num_questions} câu hỏi.")

# Sheet TN câu trả lời đúng (HL) (HSK1)
def populate_sentence_matching_shared(worksheet, data: dict):
    """
    Điền dữ liệu cho sheet 'TN câu trả lời đúng (HL) (HSK1)'.
    Key JSON: sentence_matching_questions
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    questions = data.get('questions', [])
    if not questions:
        print("   ⚠️ Cảnh báo: Không có câu hỏi nào để điền.")
        return

    start_row = 2
    num_questions = len(questions)
    end_row = start_row + num_questions - 1

    # Cột B: Học liệu chung (merge)
    shared_material = data.get('shared_material', [])
    shared_material_lines = [""]
    for item in shared_material:
        letter = item.get('option_letter', '')
        chinese = item.get('chinese_text', '')
        pinyin = item.get('pinyin', '')
        # Định dạng thụt lề cho pinyin
        shared_material_lines.append(f"{letter}. {chinese}\n   {pinyin}")
    
    shared_material_str = "\n".join(shared_material_lines)
    
    if num_questions > 0:
        worksheet.merge_cells(f'B{start_row}:B{end_row}')
        top_cell = worksheet[f'B{start_row}']
        top_cell.value = shared_material_str
        top_cell.alignment = Alignment(vertical='top', wrap_text=True, indent=0)

    # Cột F và H: Câu hỏi và đáp án
    for i, question in enumerate(questions):
        current_row = start_row + i
        # Cột D: Độ khó
        worksheet[f'D{current_row}'] = question.get('do_kho', 'NB')
        # Cột F: Câu hỏi
        chinese = question.get('question_text_chinese', '')
        pinyin = question.get('pinyin', '')
        worksheet[f'F{current_row}'] = f"{chinese} （ ）\n{pinyin}"
        # Cột G: Số lượng lựa chọn (cố định là 4)
        worksheet[f'G{current_row}'] = 4
        # Cột H: Đáp án
        # tạo map label đáp án với số thứ tự
        map_labels = {"A": 1, "B": 2, "C": 3, "D": 4, "E": 5}
        answer_label = question.get('correct_answer', '')
        if answer_label in map_labels:
            answer_index = map_labels[answer_label]
            worksheet[f'H{current_row}'] = f"{answer_index}"

    print(f"   ✅ Hoàn thành điền học liệu và {num_questions} câu hỏi.")

# Sheet TN chọn từ đúng (HL) (HSK1)
def populate_word_fill_in_shared(worksheet, data: dict):
    """
    Điền dữ liệu cho sheet 'TN chọn từ đúng (HL) (HSK1)'.
    Key JSON: word_fill_in_questions
    (Nâng cấp để xử lý cả câu đơn và hội thoại)
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    questions = data.get('questions', [])
    if not questions:
        print("   ⚠️ Cảnh báo: Không có câu hỏi nào để điền.")
        return

    start_row = 2
    num_questions = len(questions)
    end_row = start_row + num_questions - 1

    # Cột B: Học liệu chung (merge) - Logic này không đổi
    shared_material = data.get('shared_material', [])
    shared_material_lines = [""]
    for item in shared_material:
        letter = item.get('option_letter', '')
        chinese = item.get('chinese_word', '')
        pinyin = item.get('pinyin', '')
        shared_material_lines.append(f"{letter}. {chinese}\n   {pinyin}")

    shared_material_str = "\n".join(shared_material_lines)

    if num_questions > 0:
        worksheet.merge_cells(f'B{start_row}:B{end_row}')
        top_cell = worksheet[f'B{start_row}']
        top_cell.value = shared_material_str
        top_cell.alignment = Alignment(vertical='top', wrap_text=True, indent=0)

    # Cột F và H: Câu hỏi và đáp án (Logic được nâng cấp)
    for i, question in enumerate(questions):
        current_row = start_row + i
        
        # Cột D: Độ khó
        worksheet[f'D{current_row}'] = question.get('do_kho', 'NB')
        
        # Cột F: Câu hỏi (đã có thể xử lý hội thoại)
        question_lines_formatted = []
        for line in question.get('lines', []):
            speaker = line.get('speaker', '')
            prefix = f"{speaker}：" if speaker else ""
            chinese = line.get('chinese_text', '')
            pinyin = line.get('pinyin', '')
            question_lines_formatted.append(f"{prefix}{chinese}\n{' ' * len(prefix)}{pinyin}")
        
        worksheet[f'F{current_row}'] = "\n".join(question_lines_formatted)

        # Cột G: Số lượng lựa chọn (cố định là 4)
        worksheet[f'G{current_row}'] = 4
        
        # Cột H: Đáp án - Logic này không đổi
        answer_label = question.get('correct_answer', '')
        if answer_label:
            # Tạo map label đáp án với số thứ tự
            map_labels = {"A": 1, "B": 2, "C": 3, "D": 4, "E": 5}
            if answer_label in map_labels:
                answer_index = map_labels[answer_label]
                worksheet[f'H{current_row}'] = f"{answer_index}"
            else:
                worksheet[f'H{current_row}'] = answer_label

    print(f"   ✅ Hoàn thành điền học liệu và {num_questions} câu hỏi.")

# Dạng 1 - HSK2: Đúng sai (img) (HSK2)
def populate_true_false_image(worksheet, data: list):
    """
    Điền dữ liệu cho sheet 'Đúng sai (img) (HSK2)'.
    Key JSON: true_false_image_questions
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    start_row = 2  # Bắt đầu điền từ hàng thứ 2
    for i, question in enumerate(data, start=start_row):
        # Cột F: Mô tả ảnh + script
        image_des = question.get('image_des', '')
        script = question.get('script', '')
        content_f = f"Ảnh: {image_des}\nScript: {script}"
        worksheet[f'F{i}'] = content_f
        
        # Cột H: Đáp án
        answer = question.get('correct_answer')
        worksheet[f'H{i}'] = f"đúng sai: {answer}"
        
    print(f"   ✅ Hoàn thành điền {len(data)} câu hỏi.")

# Dạng 3 & 4 - HSK2: TN PA đúng (HSK2) và TN PA đúng_v2 (HSK2)
def populate_reading_comprehension_dialogue(worksheet, data: list):
    """
    Điền dữ liệu cho các sheet đọc hiểu hội thoại.
    Hàm này xử lý cả hội thoại ngắn (2 lượt lời) và dài (4 lượt lời).
    Key JSON: reading_comprehension_dialogue_2_lines / reading_comprehension_dialogue_4_lines
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    start_row = 2
    for i, question in enumerate(data, start=start_row):
        # --- Cột F: Hội thoại + Câu hỏi ---
        dialogue_parts = []
        for line in question.get('dialogue', []):
            speaker = line.get('speaker', '')
            chinese = line.get('line_chinese', '')
            pinyin = line.get('line_pinyin', '')
            # Thêm thụt lề cho pinyin để dễ đọc
            dialogue_parts.append(f"{speaker}： {chinese}")
        
        query = question.get('query_chinese', '')
        dialogue_parts.append(f"问： {query}")
        
        worksheet[f'F{i}'] = "\n".join(dialogue_parts)
        
        # --- Cột G: Các lựa chọn đáp án ---
        options = question.get('answer_options', [])
        options_parts = []
        for opt in options:
            chinese = opt.get('chinese_text', '')
            pinyin = opt.get('pinyin', '')
            if pinyin:
                options_parts.append(f"{chinese}<br/>{pinyin}")
            else:
                options_parts.append(f"{chinese}")    
        worksheet[f'G{i}'] = "\n".join(options_parts)

        # --- Cột H: Đáp án đúng ---
        worksheet[f'H{i}'] = question.get('correct_answer')
        
    print(f"   ✅ Hoàn thành điền {len(data)} câu hỏi.")

# Dạng 6 - HSK2: ĐS lựa chọn (HSK2)
def populate_true_false_statement(worksheet, data: list):
    """
    Điền dữ liệu cho sheet 'ĐS lựa chọn (HSK2)'.
    Key JSON: true_false_statement_questions
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    start_row = 2
    for i, question in enumerate(data, start=start_row):
        # Cột F: Script + Statement
        script_chinese = question.get('script_chinese', '')
        script_pinyin = question.get('script_pinyin', '')
        statement_chinese = question.get('statement_chinese', '')
        statement_pinyin = question.get('statement_pinyin', '')
        
        # Thêm thụt lề cho pinyin của statement để dễ đọc
        content_f = f"{script_chinese}\n{script_pinyin}\n★ {statement_chinese}\n    {statement_pinyin}"
        worksheet[f'F{i}'] = content_f
        
        # Cột H: Đáp án
        answer = question.get('correct_answer')
        worksheet[f'H{i}'] = f"đúng sai: {answer}"

    print(f"   ✅ Hoàn thành điền {len(data)} câu hỏi.")

# Dạng 8 - HSK2: TN câu trả lời đúng (HL) (HSK2)
def populate_sentence_matching_inverted(worksheet, data: dict):
    """
    Điền dữ liệu cho sheet 'TN câu trả lời đúng (HL) (HSK2)'.
    Logic tương tự như hàm populate_sentence_matching_shared của HSK1.
    Key JSON: sentence_matching_inverted_questions
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    questions = data.get('questions', [])
    if not questions:
        print("   ⚠️ Cảnh báo: Không có câu hỏi nào để điền.")
        return

    start_row = 2
    num_questions = len(questions)
    end_row = start_row + num_questions - 1

    # Cột B: Học liệu chung (merge)
    shared_material = data.get('shared_material', [])
    shared_material_lines = [""]
    for item in shared_material:
        letter = item.get('option_letter', '')
        chinese = item.get('chinese_text', '')
        pinyin = item.get('pinyin', '')
        shared_material_lines.append(f"{letter}. {chinese}\n   {pinyin}")
    
    shared_material_str = "\n".join(shared_material_lines)
    
    if num_questions > 0:
        worksheet.merge_cells(f'B{start_row}:B{end_row}')
        top_cell = worksheet[f'B{start_row}']
        top_cell.value = shared_material_str
        top_cell.alignment = Alignment(vertical='top', wrap_text=True, indent=0)

    # Cột F và H: Câu hỏi và đáp án
    for i, question in enumerate(questions):
        current_row = start_row + i
        # Cột F: Câu hỏi
        chinese = question.get('question_text_chinese', '')
        pinyin = question.get('pinyin', '')
        # Theo ví dụ, câu hỏi có thể có dấu ( ) ở cuối
        worksheet[f'F{current_row}'] = f"{chinese} (   )\n{pinyin}"
        # Cột H: Đáp án
        # Tạo map label đáp án với số thứ tự
        map_labels = {"A": 1, "B": 2, "C": 3, "D": 4, "E": 5, "F": 6}
        answer_label = question.get('correct_answer', '')
        if answer_label in map_labels:
            answer_index = map_labels[answer_label]
            worksheet[f'H{current_row}'] = f"{answer_index}"

    print(f"   ✅ Hoàn thành điền học liệu và {num_questions} câu hỏi.")

# Dạng 2 - HSK3: ĐS nghe chọn (HSK3)  
def populate_true_false_listening_choice(worksheet, data: list):
    """
    Điền dữ liệu cho sheet 'ĐS nghe chọn (HSK3)'.
    Key JSON: true_false_listening_choice
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    start_row = 2
    for i, question in enumerate(data, start=start_row):
        # Cột F: Statement + Script
        statement = question.get('statement_chinese', '')
        script = question.get('script_chinese', '')
        
        content_f = f"{statement} （ ）\nScript: {script}"
        worksheet[f'F{i}'] = content_f
        
        # Cột H: Đáp án
        answer = question.get('correct_answer')
        worksheet[f'H{i}'] = f"đúng sai: {answer}"

    print(f"   ✅ Hoàn thành điền {len(data)} câu hỏi.")

# Dạng 7: TN PA đúng ko pinyin_v3 (HSK3)
def populate_reading_comprehension_v3(worksheet, data: list):
    """
    Điền dữ liệu cho sheet 'TN PA đúng ko pinyin_v3 (HSK3)'.
    Key JSON: reading_comprehension_no_pinyin_v3
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    start_row = 2
    for i, question in enumerate(data, start=start_row):
        # Cột F: Script + Query
        script = question.get('script_chinese', '')
        query = question.get('query_chinese', '')
        worksheet[f'F{i}'] = f"{script}\n{query}"
        
        # Cột G: Các lựa chọn đáp án
        options = question.get('answer_options', [])
        worksheet[f'G{i}'] = "\n".join(options)

        # Cột H: Đáp án đúng
        worksheet[f'H{i}'] = question.get('correct_answer')
        
    print(f"   ✅ Hoàn thành điền {len(data)} câu hỏi.")

# Dạng 8: Sắp xếp câu (HSK3)
def populate_sentence_reordering(worksheet, data: list):
    """
    Điền dữ liệu cho sheet 'Sắp xếp câu (HSK3)'.
    Key JSON: sentence_reordering
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    start_row = 2
    for i, question in enumerate(data, start=start_row):
        # Cột G: Các thành phần câu
        components = question.get('components', [])
        worksheet[f'G{i}'] = "; ".join(components)
        
        # Cột H: Đáp án đúng
        worksheet[f'H{i}'] = question.get('correct_order')
        
    print(f"   ✅ Hoàn thành điền {len(data)} câu hỏi.")

# Dạng 9: Điền từ đúng (HSK3)
def populate_word_fill_in_display_answer(worksheet, data: list):
    """
    Điền dữ liệu cho sheet 'Điền từ đúng (HSK3)'.
    Key JSON: word_fill_in_display_answer
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    start_row = 2
    for i, question in enumerate(data, start=start_row):
        # Ghép các thành phần lại thành câu hoàn chỉnh
        part1 = question.get('sentence_part_1', '')
        word = question.get('correct_word', '')
        pinyin = question.get('correct_word_pinyin', '')
        part2 = question.get('sentence_part_2', '')

        full_sentence = f"{part1} [{word}] ({pinyin}) {part2}"
        
        # Cột G: Điền câu hoàn chỉnh
        worksheet[f'G{i}'] = full_sentence
        
    print(f"   ✅ Hoàn thành điền {len(data)} câu hỏi.")

# HÀM TÁI SỬ DỤNG cho Dạng 4 (TN Nghe hiểu) và Dạng 8 (TN Đọc hiểu ngắn)
def populate_passage_with_multiple_questions(worksheet, data: list):
    """
    Điền dữ liệu cho các dạng bài có 1 học liệu chung cho nhiều câu hỏi.
    Hàm này sẽ merge ô ở cột B và điền dữ liệu cho các câu hỏi tương ứng.
    Key JSON: listening_comprehension / reading_comprehension_short_passage
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    current_row = 2  # Bắt đầu từ hàng 2

    for passage_data in data:
        questions = passage_data.get('questions', [])
        num_questions = len(questions)
        
        if num_questions == 0:
            continue

        start_row_for_merge = current_row
        end_row_for_merge = current_row + num_questions - 1

        # --- Cột B: Merge và điền học liệu ---
        merge_range = f'B{start_row_for_merge}:B{end_row_for_merge}'
        worksheet.merge_cells(merge_range)
        top_left_cell = worksheet[f'B{start_row_for_merge}']
        top_left_cell.value = passage_data.get('script_chinese', '')
        top_left_cell.alignment = Alignment(vertical='top', wrap_text=True)

        # --- Cột F, G, H: Điền từng câu hỏi ---
        for question in questions:
            worksheet[f'F{current_row}'] = question.get('query_chinese', '')
            
            options = question.get('answer_options', [])
            worksheet[f'G{current_row}'] = "\n".join(options)
            
            worksheet[f'H{current_row}'] = question.get('correct_answer')
            
            current_row += 1 # Tăng số hàng cho câu hỏi tiếp theo

    print(f"   ✅ Hoàn thành điền {len(data)} học liệu.")

# Hàm cho Dạng 6: Sắp xếp các câu (HSK4)
def populate_sentence_sequence_reordering(worksheet, data: list):
    """
    Điền dữ liệu cho sheet 'Sắp xếp các câu (HSK4)'.
    Key JSON: sentence_sequence_reordering
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    start_row = 2
    for i, question in enumerate(data, start=start_row):
        
        components = question.get('components', [])
        correct_order = question.get('correct_order', '')
        
        # Ghép các thành phần và đáp án thành một chuỗi duy nhất
        lines = []
        for comp in components:
            label = comp.get('label', '')
            text = comp.get('text', '')
            lines.append(f"{label}. {text}")
            
        lines.append(f"[{correct_order}]")
        
        # Cột G: Điền toàn bộ nội dung
        worksheet[f'G{i}'] = "\n".join(lines)
        
    print(f"   ✅ Hoàn thành điền {len(data)} câu hỏi.")

def populate_image_with_word_sentence_creation(worksheet, data: list):
    """
    Điền dữ liệu cho sheet 'Ảnh với từ (img) (HSK4)'.
    Key JSON: image_with_word_sentence_creation
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    start_row = 2
    for i, question in enumerate(data, start=start_row):
        # Lấy dữ liệu từ JSON
        image_des = question.get('image_description', '')
        vocab = question.get('vocabulary_word', '')
        sample_sentence = question.get('sample_sentence', '') # Đáp án mẫu

        # Xây dựng nội dung cho cột E (Câu hỏi cho người học)
        content_e = f"Ảnh: {image_des}\nTừ vựng: {vocab}"
        
        # Quy tắc đặc biệt: Thêm đề bài chung cho câu đầu tiên
        if i == start_row:
            header = "看图，用词造句。\n\n" # Đề bài: Nhìn hình, dùng từ đặt câu.
            content_e = header + content_e
            
        # Điền dữ liệu vào sheet
        worksheet[f'E{i}'] = content_e
        worksheet[f'H{i}'] = sample_sentence # Cột H giờ là câu trả lời mẫu

    print(f"   ✅ Hoàn thành điền {len(data)} câu hỏi.")

# Hàm cho HSK5
def populate_passage_cloze(worksheet, data: dict):
    """
    Điền dữ liệu cho sheet 'Điền từ đoạn văn (HSK5)'.
    Key JSON: passage_cloze
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    questions = data.get('questions', [])
    num_questions = len(questions)

    if num_questions == 0:
        print("   ⚠️ Cảnh báo: Không có câu hỏi nào để điền.")
        return

    start_row = 2
    end_row = start_row + num_questions - 1

    # Cột B: Merge ô và điền học liệu
    merge_range = f'B{start_row}:B{end_row}'
    worksheet.merge_cells(merge_range)
    top_left_cell = worksheet[f'B{start_row}']
    top_left_cell.value = data.get('shared_material', '')
    top_left_cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Cột G và H: Điền các bộ phương án và đáp án
    for i, question in enumerate(questions, start=start_row):
        # Cột G: Phương án trả lời
        options = question.get('answer_options', [])
        worksheet[f'G{i}'] = "\n".join(options)
        
        # Cột H: Đáp án
        worksheet[f'H{i}'] = question.get('correct_answer')

    print(f"   ✅ Hoàn thành điền học liệu và {num_questions} câu hỏi.")

def populate_main_idea_comprehension(worksheet, data: list):
    """
    Điền dữ liệu cho sheet 'Chọn chủ đề đoạn văn (HSK5)'.
    Key JSON: main_idea_comprehension
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    start_row = 2
    for i, question in enumerate(data, start=start_row):
        # Cột F: Đoạn văn
        worksheet[f'F{i}'] = question.get('passage_chinese', '')
        
        # Cột G: Các lựa chọn đáp án
        options = question.get('answer_options', [])
        worksheet[f'G{i}'] = "\n".join(options)

        # Cột H: Đáp án đúng
        worksheet[f'H{i}'] = question.get('correct_answer')
        
    print(f"   ✅ Hoàn thành điền {len(data)} câu hỏi.")

def populate_long_passage_comprehension(worksheet, data: dict):
    """
    Điền dữ liệu cho sheet 'TN Đọc hiểu (img) (HSK5)'.
    Key JSON: long_passage_comprehension
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    questions = data.get('questions', [])
    num_questions = len(questions)

    if num_questions == 0:
        print("   ⚠️ Cảnh báo: Không có câu hỏi nào để điền.")
        return

    start_row = 2
    end_row = start_row + num_questions - 1
    
    # --- Cột B: Merge ô và điền học liệu + mô tả ảnh ---
    passage = data.get('shared_passage', '')
    image_des = data.get('image_description', '')
    
    full_material = f"{passage}\nẢnh : {image_des}"
    
    merge_range = f'B{start_row}:B{end_row}'
    worksheet.merge_cells(merge_range)
    top_left_cell = worksheet[f'B{start_row}']
    top_left_cell.value = full_material
    top_left_cell.alignment = Alignment(vertical='top', wrap_text=True)

    # --- Cột F, G, H: Điền các câu hỏi, phương án, và đáp án ---
    for i, question in enumerate(questions, start=start_row):
        worksheet[f'F{i}'] = question.get('query_chinese', '')
        
        options = question.get('answer_options', [])
        worksheet[f'G{i}'] = "\n".join(options)
        
        worksheet[f'H{i}'] = question.get('correct_answer')

    print(f"   ✅ Hoàn thành điền học liệu và {num_questions} câu hỏi.")

def populate_writing_from_keywords(worksheet, data: list):
    """
    Điền dữ liệu cho sheet 'Viết dựa vào từ (HSK5)'.
    Key JSON: writing_from_keywords
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    start_row = 2
    for i, question in enumerate(data, start=start_row):
        # Lấy dữ liệu
        keywords = question.get('keywords', [])
        paragraph = question.get('sample_paragraph_marked', '')
        translation = question.get('translation', '')
        
        # Cột E: Đề bài và từ khóa
        prompt_text = "请结合下列词语(要全部使用,顺序不分先后),写一篇80字左右的短文。"
        keywords_text = "\t".join(keywords) # Dùng tab để có khoảng cách đẹp
        content_e = f"{prompt_text}\n\n{keywords_text}"
        worksheet[f'E{i}'] = content_e
        
        # Cột I: Đoạn văn mẫu và bản dịch
        content_i = f"{paragraph}\n\nTạm dịch:\n{translation}"
        worksheet[f'I{i}'] = content_i

    print(f"   ✅ Hoàn thành điền {len(data)} câu hỏi.")

def populate_writing_from_image(worksheet, data: list):
    """
    Điền dữ liệu cho sheet 'Viết dựa vào ảnh (img) (HSK5)'.
    Key JSON: writing_from_image
    """
    print(f"   -> Đang điền dữ liệu vào sheet: {worksheet.title}")
    start_row = 2
    for i, question in enumerate(data, start=start_row):
        # Lấy dữ liệu
        image_des = question.get('image_description', '')
        paragraph = question.get('sample_paragraph', '')
        translation = question.get('translation', '')

        # Cột E: Đề bài và mô tả ảnh
        prompt_text = "请结合这张图片写一篇80字左右的短文。"
        content_e = f"Ảnh: {image_des}\n{prompt_text}"
        worksheet[f'E{i}'] = content_e

        # Cột I: Đoạn văn mẫu, bản dịch và ghi chú AICham
        content_i = f"{paragraph}\n\nTạm dịch:\n{translation}\nAICham: {image_des}"
        worksheet[f'I{i}'] = content_i

    print(f"   ✅ Hoàn thành điền {len(data)} câu hỏi.")