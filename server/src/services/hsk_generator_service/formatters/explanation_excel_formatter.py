# hsk_core/formatters/explanation_excel_formatter.py
import re
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# --- CÁC HÀM TIỆN ÍCH CHO AUTO-SIZING ---
def calculate_cell_dimensions(text: str, font_size=11):
    """
    Tính toán chiều rộng và chiều cao cần thiết cho ô dựa trên nội dung text.
    
    Args:
        text (str): Nội dung text cần tính toán
        font_size (int): Kích thước font (mặc định 11)
    
    Returns:
        tuple: (width, height) - chiều rộng và số dòng cần thiết
    """
    if not text:
        return 15, 1
    
    # Xử lý text từ CellRichText hoặc string thông thường
    if hasattr(text, '__iter__') and not isinstance(text, str):
        # Nếu là CellRichText, chuyển về string để tính toán
        plain_text = ""
        for part in text:
            if isinstance(part, TextBlock):
                plain_text += part.text
            else:
                plain_text += str(part)
        text = plain_text
    
    lines = str(text).split('\n')
    max_line_length = max(len(line) for line in lines) if lines else 0
    num_lines = len(lines)
    
    # Tính chiều rộng: khoảng 1.2 đơn vị Excel cho mỗi ký tự (tùy chỉnh theo font)
    # Đối với tiếng Việt và tiếng Trung, cần thêm không gian
    char_width = 1.3
    estimated_width = max_line_length * char_width
    
    # Đặt giới hạn tối thiểu và tối đa
    min_width = 20
    max_width = 100
    final_width = max(min_width, min(estimated_width, max_width))
    
    # Tính số dòng thực tế cần thiết (không tăng quá nhiều)
    actual_lines = num_lines
    if estimated_width > max_width:
        # Chỉ tăng nhẹ số dòng khi text quá dài
        wrap_factor = estimated_width / max_width
        additional_lines = max(0, int((wrap_factor - 1) * 0.5))  # Giảm hệ số wrap
        actual_lines = num_lines + additional_lines
    
    return final_width, actual_lines

def auto_size_cell(worksheet, cell, content):
    """
    Tự động điều chỉnh kích thước ô theo nội dung.
    
    Args:
        worksheet: Worksheet object của openpyxl
        cell: Cell object cần điều chỉnh
        content: Nội dung đã được ghi vào cell
    """
    try:
        # Tính toán kích thước cần thiết
        width, num_lines = calculate_cell_dimensions(content)
        
        # Lấy chỉ số cột
        col_letter = get_column_letter(cell.column)
        
        # Điều chỉnh chiều rộng cột (chỉ tăng, không giảm)
        current_width = worksheet.column_dimensions[col_letter].width or 15
        if width > current_width:
            worksheet.column_dimensions[col_letter].width = width
        
        # Điều chỉnh chiều cao hàng
        row_height = max(15, num_lines * 15)  # 15 points per line
        current_height = worksheet.row_dimensions[cell.row].height or 15
        if row_height > current_height:
            worksheet.row_dimensions[cell.row].height = row_height
            
    except Exception as e:
        print(f"Cảnh báo: Không thể auto-size cell {cell.coordinate}: {e}")

# --- CÁC HÀM XÂY DỰNG RICH TEXT CHUYÊN BIỆT ---
def format_simple_context_rich_text(explanation_json: dict) -> CellRichText:
    """Xây dựng CellRichText cho các dạng câu hỏi có ngữ cảnh đơn giản."""
    bold_font = InlineFont(b=True)
    italic_font = InlineFont(i=True)
    rich_text = CellRichText()

    # 1. Phân tích
    rich_text.append(explanation_json.get('analysis_paragraph', '') + '\n\n')

    # 2. Phụ đề
    context = explanation_json.get('context_block', {})
    rich_text.append(TextBlock(bold_font, "Phụ đề:\n"))
    rich_text.append(f"{context.get('chinese_text', '')}\n")
    rich_text.append(f"{context.get('pinyin', '')}\n\n")

    # 3. Tạm dịch
    translation = explanation_json.get('translation_block', {})
    rich_text.append(TextBlock(bold_font, "Tạm dịch: "))
    rich_text.append(TextBlock(italic_font, translation.get('vietnamese_text', '')))
    
    return rich_text

def format_dialogue_rich_text(explanation_json: dict) -> CellRichText:
    """Xây dựng CellRichText cho các dạng câu hỏi có hội thoại."""
    bold_font = InlineFont(b=True)
    italic_font = InlineFont(i=True)
    rich_text = CellRichText()

    # 1. Phân tích
    rich_text.append(explanation_json.get('analysis_paragraph', '') + '\n\n')
    
    # 2. Phụ đề (hội thoại)
    context = explanation_json.get('context_block', {})
    dialogue = context.get('dialogue', [])
    rich_text.append(TextBlock(bold_font, "Phụ đề:\n"))
    for turn in dialogue:
        speaker = turn.get('speaker', '')
        speaker_prefix = f"**{speaker}：**" if speaker else ""
        rich_text.append(f"{speaker_prefix}{turn.get('chinese_text', '')}\n")
        rich_text.append(f"{' ' * len(speaker_prefix)}{turn.get('pinyin', '')}\n")
    rich_text.append('\n')

    # 3. Tạm dịch
    translation = explanation_json.get('translation_block', {})
    rich_text.append(TextBlock(bold_font, "Tạm dịch:\n"))
    rich_text.append(TextBlock(italic_font, translation.get('vietnamese_text', '')))

    return rich_text

def format_word_fill_rich_text(explanation_json: dict) -> CellRichText:
    """Xây dựng CellRichText cho dạng Điền từ, có xử lý highlight chữ Hán."""
    bold_font = InlineFont(b=True)
    italic_font = InlineFont(i=True)
    blue_font = InlineFont(color="0000FF") # Màu xanh dương cho dễ nhìn
    rich_text = CellRichText()
    
    # 1. Phân tích
    rich_text.append(explanation_json.get('analysis_paragraph', '') + '\n\n')

    # 2. Phụ đề (với highlight)
    context = explanation_json.get('context_block', {})
    script_lines = context.get('script_lines', [])
    rich_text.append(TextBlock(bold_font, "Phụ đề:\n"))
    
    chinese_pattern = r'(__[^_]+__)' # Pattern để tìm __text__
    for line in script_lines:
        speaker = line.get('speaker', '')
        prefix = f"**{speaker}：**" if speaker else ""
        chinese_text = line.get('chinese_text', '')
        pinyin = line.get('pinyin', '')
        
        rich_text.append(prefix)
        # Tách chuỗi để highlight
        parts = re.split(chinese_pattern, chinese_text)
        for i, part in enumerate(parts):
            if i % 2 == 1: # Phần nằm giữa cặp __
                rich_text.append(TextBlock(blue_font, part[2:-2])) # Bỏ đi 2 dấu __
            else:
                rich_text.append(part)
        rich_text.append('\n')
        rich_text.append(f"{' ' * len(prefix)}{pinyin}\n")
    rich_text.append('\n')

    # 3. Tạm dịch
    translation = explanation_json.get('translation_block', {})
    vietnamese_text = translation.get('vietnamese_text', '')
    if vietnamese_text.startswith("**Nam:**") or vietnamese_text.startswith("**Nữ:**"):
        rich_text.append(TextBlock(bold_font, "Tạm dịch:\n"))
    else:
        rich_text.append(TextBlock(bold_font, "Tạm dịch: "))
    rich_text.append(TextBlock(italic_font, vietnamese_text))
    
    return rich_text

def apply_ds_vocab_rich_text_formatting(text: str) -> CellRichText:
    """
    Áp dụng định dạng rich text cho dạng DS vocab với các quy tắc:
    - Bôi đậm: "Phụ đề:", "Tạm dịch:"
    - In nghiêng: Text sau "Tạm dịch:"
    """
    # Define fonts
    bold_font = InlineFont(b=True)
    italic_font = InlineFont(i=True)
    
    rich_text = CellRichText()
    
    # Define keywords để bôi đậm
    bold_keywords = ["Phụ đề:", "Tạm dịch:"]
    
    # Tạo pattern regex
    escaped_keywords = [re.escape(kw) for kw in bold_keywords]
    pattern = '(' + '|'.join(escaped_keywords) + ')'
    
    # Split text theo keywords, giữ lại các keywords
    parts = re.split(pattern, text)
    
    in_tam_dich_section = False
    
    for part in parts:
        if not part:  # Skip empty parts
            continue
            
        if part in bold_keywords:
            rich_text.append(TextBlock(bold_font, part))
            if part == "Tạm dịch:":
                in_tam_dich_section = True
        else:
            # Normal text
            if in_tam_dich_section:
                rich_text.append(TextBlock(italic_font, part))
            else:
                rich_text.append(part)
    
    return rich_text

# apply rich text formatting cho sheet "TN PA đúng (img) (HL) (HSK1)"
def build_plain_text_with_markers(explanation_json: dict) -> str:
    text_parts = []

    # 1) analysis...
    analysis = explanation_json.get('analysis_paragraph', '')
    if analysis:
        text_parts.append(analysis)
        text_parts.append('\n\n')

    # Thêm option ở dạng 3,4 của HSK2
    options = explanation_json.get('options_list', [])
    if options:
        for opt in options:
            letter = opt.get('letter', '')
            chinese = opt.get('chinese_text', '')
            translation = opt.get('translation', '')
            pinyin = opt.get('pinyin', '')
            # Dùng marker <<i>> cho phần dịch nghĩa
            text_parts.append(f"{letter}. {chinese} <<i>>({translation})<</i>>\n    {pinyin}\n")
        text_parts.append('\n')

    # 2) phụ đề...
    context = explanation_json.get('context_block', {}) or {}
    dialogue = context.get('dialogue', []) or []
    if dialogue:
        text_parts.append('**Phụ đề**:\n')
        for i, turn in enumerate(dialogue):
            speaker = turn.get('speaker', '') or ''
            chinese = turn.get('chinese_text', '') or ''
            pinyin  = turn.get('pinyin', '') or ''
            if speaker:
                text_parts.append(f'**{speaker}**：')
            if chinese:
                text_parts.append(chinese)
            text_parts.append('\n')
            if pinyin:
                pinyin_indent = ' ' * len(f"{speaker}：") if speaker else ''
                text_parts.append(f"{pinyin_indent}{pinyin}")
            if i < len(dialogue) - 1:
                text_parts.append('\n')
        text_parts.append('\n\n')

    # 3) Tạm dịch (sửa tại đây)
    translation = explanation_json.get('translation_block', {}) or {}
    vi_turns = translation.get('vietnamese_dialogue', []) or []
    if vi_turns:
        text_parts.append('**Tạm dịch**:\n')
        for i, tvi in enumerate(vi_turns):
            sp = tvi.get('speaker_vi', '') or ''
            dv = tvi.get('dialogue_vi', '') or ''
            if sp:
                # speaker Việt: đậm + nghiêng
                text_parts.append(f'<<bi>>{sp}<</bi>>: ')
            if dv:
                # NHÉT \n VÀO BÊN TRONG marker italic (trừ dòng cuối)
                if i < len(vi_turns) - 1:
                    text_parts.append(f'<<i>>{dv}\n<</i>>')
                else:
                    text_parts.append(f'<<i>>{dv}<</i>>')

    return ''.join(text_parts)

# Áp dụng regex để build CellRichText từ plain text với markers
def apply_regex_formatting(plain_text: str) -> CellRichText:
    """
    Parse các marker và build CellRichText.
    Thứ tự ưu tiên: <<bi>>...<</bi>> -> <<i>>...<</i>> -> **...**
    """
    bold_font        = InlineFont(b=True)
    italic_font      = InlineFont(i=True)
    bold_italic_font = InlineFont(b=True, i=True)

    rich = CellRichText()

    # Combined token regex (non-greedy, cho phép xuống dòng)
    token_re = re.compile(r'(?s)(<<bi>>.*?<<\/bi>>|<<i>>.*?<<\/i>>|\*\*[^*]+\*\*)')

    pos = 0
    for m in token_re.finditer(plain_text):
        # 1) text thường trước token
        if m.start() > pos:
            rich.append(plain_text[pos:m.start()])

        tok = m.group(0)

        if tok.startswith('<<bi>>'):
            content = tok[len('<<bi>>'):-len('<</bi>>')]
            rich.append(TextBlock(bold_italic_font, content))
        elif tok.startswith('<<i>>'):
            content = tok[len('<<i>>'):-len('<</i>>')]
            rich.append(TextBlock(italic_font, content))
        else:
            # **...**
            content = tok[2:-2]
            rich.append(TextBlock(bold_font, content))

        pos = m.end()

    # 2) phần còn lại sau token cuối
    if pos < len(plain_text):
        rich.append(plain_text[pos:])

    return rich

# render rich text cho sheet "TN PA đúng (img) (HL) (HSK1)"
def format_shared_image_comprehension_rich_text(explanation_json: dict) -> CellRichText:
    """
    Xây dựng CellRichText cho sheet "TN PA đúng (img) (HL) (HSK1)".
    Sử dụng phương pháp tạo plain text với markers rồi áp dụng regex formatting.
    """
    
    # 1. Tạo plain text với markers
    plain_text_with_markers = build_plain_text_with_markers(explanation_json)
    
    # 2. Áp dụng định dạng bằng regex
    rich_text = apply_regex_formatting(plain_text_with_markers)
    
    return rich_text

# apply rich text formatting cho sheet "TN PA đúng (HSK1)"
def apply_rich_text_formatting(text: str) -> CellRichText:
    bold_font = InlineFont(b=True)
    italic_font = InlineFont(i=True)
    bold_italic_font = InlineFont(b=True, i=True)

    rich_text = CellRichText()

    bold_keywords = ["Tạm dịch:", "Phụ đề:"]
    bold_italic_keywords = ["Câu hỏi:"]
    all_keywords = bold_keywords + bold_italic_keywords

    # Tách theo heading (giữ lại heading)
    escaped_keywords = [re.escape(kw) for kw in sorted(all_keywords, key=len, reverse=True)]
    pattern = '(' + '|'.join(escaped_keywords) + ')'
    parts = re.split(pattern, text)

    # NEW: bắt token **...** và <<i>>...<<\/i>> (non-greedy, đa dòng)
    bold_token_re = re.compile(r'(?s)(<<bi>>.*?<<\/bi>>|<<i>>.*?<<\/i>>|\*\*[^*]+\*\*)')

    def emit_with_bold_tokens(chunk: str, italic_ctx: bool):
        pos = 0
        for m in bold_token_re.finditer(chunk):
            if m.start() > pos:
                normal = chunk[pos:m.start()]
                if italic_ctx:
                    rich_text.append(TextBlock(italic_font, normal))
                else:
                    rich_text.append(normal)
            
            # Xử lý các token
            token = m.group(0)
            if token.startswith('<<i>>') and token.endswith('<</i>>'):
                # Token italic đơn thuần
                content = token[5:-6]  # Bỏ <<i>> và <</i>>
                rich_text.append(TextBlock(italic_font, content))
            elif token.startswith('<<bi>>') and token.endswith('<</bi>>'):
                # Token bold italic
                content = token[6:-7]  # Bỏ <<bi>> và <</bi>>
                rich_text.append(TextBlock(bold_italic_font, content))
            elif token.startswith('**') and token.endswith('**'):
                # Token bold (ví dụ **问：**)
                content = token[2:-2]
                rich_text.append(TextBlock(bold_font, content))
            
            pos = m.end()
        
        if pos < len(chunk):
            tail = chunk[pos:]
            if italic_ctx:
                rich_text.append(TextBlock(italic_font, tail))
            else:
                rich_text.append(tail)

    in_tam_dich_section = False

    for part in parts:
        if not part:
            continue
        if part in bold_keywords:
            rich_text.append(TextBlock(bold_font, part))
            in_tam_dich_section = (part == "Tạm dịch:")
        elif part in bold_italic_keywords:
            rich_text.append(TextBlock(bold_italic_font, part))
            # KHÔNG reset in_tam_dich_section - giữ nguyên để phần sau "Câu hỏi:" vẫn italic
        else:
            # Không phải heading -> xử lý **...** và <<i>>...<<\/i>> tokens
            emit_with_bold_tokens(part, italic_ctx=in_tam_dich_section)

    return rich_text

# render lời giải cho sheet "TN PA đúng (HSK1)"
def render_reading_comp_explanation(cell, explanation_json: dict):
    """
    Ghi lời giải cho dạng Đọc hiểu (TN PA đúng) vào ô Excel với rich text formatting.
    """
    # Initialize cell value as a string
    cell_value = ""
    
    # 1. Đoạn phân tích
    analysis = explanation_json.get('analysis_paragraph', '')
    cell_value += analysis + '\n\n'
    
    # 2. Danh sách lựa chọn
    options = explanation_json.get('options_list', [])
    for opt in options:
        letter = opt.get('letter', '')
        chinese = opt.get('chinese_text', '')
        translation = opt.get('translation', '')
        pinyin = opt.get('pinyin', '')
        cell_value += f"{letter}. {chinese} <<i>>({translation})<</i>>\n"
        cell_value += f"{pinyin}\n"
    cell_value += '\n'

    # 3. Khối Phụ đề
    context = explanation_json.get('context_block', {})
    script_lines = context.get('script_lines', [])
    cell_value += "Phụ đề:\n"
    for line in script_lines:
        label = line.get('line_label', '')
        prefix = f"**{label}**：" if label else ""
        cell_value += f"{prefix}{line.get('chinese_text', '')}\n"
        cell_value += f"{' ' * len(prefix)}{line.get('pinyin', '')}\n"
    cell_value += '\n'

    # 4. Khối Tạm dịch
    translation = explanation_json.get('translation_block', {})
    cell_value += "Tạm dịch:\n"
    cell_value += f"{translation.get('context_vietnamese', '')}\n"
    cell_value += "Câu hỏi: "
    cell_value += f"{translation.get('query_vietnamese', '')}"

    print(cell_value)

    # Apply rich text formatting
    cell.value = apply_rich_text_formatting(cell_value)
    cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Auto-size the cell
    auto_size_cell(cell.parent, cell, cell_value)

_CJK_SPEAKER_SET = {"男", "女", "问", "答", "A", "B"}
_CJK_LABEL_RE    = re.compile(r'^[\u3400-\u4DBF\u4E00-\u9FFF]{1,3}$')  # 1–3 ký tự CJK
_VI_SPEAKER_RE   = re.compile(r'^\*{0,2}(Nam|Nữ|A|B)\*{0,2}\s*:\s*')

def is_dialogue_style_word_fill(expl: dict) -> bool:
    ctx = expl.get("context_block", {}) or {}
    sl  = ctx.get("script_lines", []) or []
    if sl:
        speakers = [(line.get("speaker") or "").strip() for line in sl]
        if any((s in _CJK_SPEAKER_SET) or _CJK_LABEL_RE.match(s) for s in speakers if s):
            return True

    vt = ((expl.get("translation_block") or {}).get("vietnamese_text") or "")
    if any(_VI_SPEAKER_RE.match(line.strip()) for line in vt.splitlines()):
        return True

    return False

def normalize_word_fill_to_dialogue_schema(expl: dict) -> dict:
    sl = (expl.get("context_block") or {}).get("script_lines", []) or []
    dialogue = []
    for line in sl:
        dialogue.append({
            "speaker":      (line.get("speaker") or "").strip(),
            "chinese_text": (line.get("chinese_text") or ""),
            "pinyin":       (line.get("pinyin") or "")
        })

    vt = ((expl.get("translation_block") or {}).get("vietnamese_text") or "").strip()
    vi_dialogue = []
    if vt:
        for raw in [s.strip() for s in vt.splitlines() if s.strip()]:
            m = re.match(r'^\*{0,2}(Nam|Nữ)\*{0,2}\s*:\s*(.+)$', raw)
            if m:
                vi_dialogue.append({"speaker_vi": m.group(1), "dialogue_vi": m.group(2)})
        if not vi_dialogue:
            vi_dialogue.append({"speaker_vi": "", "dialogue_vi": vt})

    return {
        "analysis_paragraph": expl.get("analysis_paragraph", ""),
        "context_block": {"dialogue": dialogue},
        "translation_block": {"vietnamese_dialogue": vi_dialogue}
    }

def apply_chinese_highlight_over_markers(plain_text: str) -> CellRichText:
    green_font = InlineFont(color="FF00AA00")  # ARGB 8 ký tự
    rich = CellRichText()
    token_re = re.compile(r'__([\u3400-\u4DBF\u4E00-\u9FFF]+?)__')

    pos = 0
    for m in token_re.finditer(plain_text):
        if m.start() > pos:
            for run in apply_regex_formatting(plain_text[pos:m.start()]):
                rich.append(run)
        rich.append(TextBlock(green_font, m.group(1)))
        pos = m.end()
    if pos < len(plain_text):
        for run in apply_regex_formatting(plain_text[pos:]):
            rich.append(run)
    return rich

def apply_rich_text_formatting_with_chinese(text: str) -> CellRichText:
    green_font = InlineFont(color="FF00AA00")  # ARGB
    rich = CellRichText()
    token_re = re.compile(r'__([\u3400-\u4DBF\u4E00-\u9FFF]+?)__')

    pos = 0
    for m in token_re.finditer(text):
        if m.start() > pos:
            for run in apply_rich_text_formatting(text[pos:m.start()]):  # hàm bạn đã patch để hiểu **…**
                rich.append(run)
        rich.append(TextBlock(green_font, m.group(1)))
        pos = m.end()

    if pos < len(text):
        for run in apply_rich_text_formatting(text[pos:]):
            rich.append(run)
    return rich

def render_word_fill_explanation(cell, explanation_json: dict):
    # Nếu là thoại -> dùng pipeline giống sheet ảnh (và vẫn highlight __漢字__)
    if is_dialogue_style_word_fill(explanation_json):
        normalized = normalize_word_fill_to_dialogue_schema(explanation_json)
        plain      = build_plain_text_with_markers(normalized)     # tạo **Phụ đề**/**Tạm dịch**, <<bi>> label
        rich       = apply_chinese_highlight_over_markers(plain)   # parse markers + overlay __漢字__
        cell.value = rich
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        auto_size_cell(cell.parent, cell, plain_text_from_rich_text(rich))
        return

    # Không phải thoại -> luồng cũ (nhưng dùng ASCII ":" và thụt pinyin cố định)
    cell_value = ""

    analysis = explanation_json.get('analysis_paragraph', '')
    cell_value += analysis + '\n\n'

    context = explanation_json.get('context_block', {})
    script_lines = context.get('script_lines', [])
    cell_value += "Phụ đề:\n"
    for line in script_lines:
        speaker = line.get('speaker', '')
        prefix  = f"**{speaker}:** " if speaker else ""   # dùng ":" ASCII
        chinese = line.get('chinese_text', '')
        pinyin  = line.get('pinyin', '')
        cell_value += f"{prefix}{chinese}\n"
        cell_value += f"  {pinyin}\n"                     # thụt cố định 2 space
    cell_value += '\n'

    tr = explanation_json.get('translation_block', {})
    vt = tr.get('vietnamese_text', '') or ''
    if vt.startswith("**Nam:**") or vt.startswith("**Nữ:**"):
        cell_value += "Tạm dịch:\n"
    else:
        cell_value += "Tạm dịch:\n"
    cell_value += vt

    cell.value = apply_rich_text_formatting_with_chinese(cell_value)
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    auto_size_cell(cell.parent, cell, cell_value)

# Render lời giải cho sheet ĐS (img) HSK1 và ĐS Ko phụ đề (img) HSK1
def render_ds_vocab_explanation(worksheet, sheet_name):
    """
    Định dạng lại nội dung có sẵn trong cột I của sheet (bỏ qua hàng tiêu đề).
    Bôi đậm: "Phụ đề:", "Tạm dịch:"
    In nghiêng: phần sau "Tạm dịch:"
    """
    print(f"   - Đang định dạng sheet: {sheet_name}")
    
    try:
        # Lấy số hàng có dữ liệu trong cột I
        max_row = worksheet.max_row
        
        # Duyệt từ hàng 2 (bỏ qua tiêu đề) đến hàng cuối
        for row_num in range(2, max_row + 1):
            cell = worksheet[f'I{row_num}']
            
            # Chỉ xử lý nếu ô có nội dung
            if cell.value:
                # Lấy nội dung text gốc
                original_text = str(cell.value)
                
                # Áp dụng định dạng rich text
                formatted_content = apply_ds_vocab_rich_text_formatting(original_text)
                cell.value = formatted_content
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Auto-size the cell
                auto_size_cell(worksheet, cell, original_text)
                
        print(f"   - ✅ Đã định dạng xong sheet: {sheet_name}")
        
    except Exception as e:
        print(f"   - ❌ Lỗi khi định dạng sheet {sheet_name}: {e}")

def plain_text_from_rich_text(rt: CellRichText) -> str:
    """Lấy chuỗi văn bản thô từ đối tượng CellRichText để tính toán kích thước."""
    return "".join(str(part.text if hasattr(part, 'text') else part) for part in rt)

def render_simple_context_explanation(cell, explanation_json: dict):
    """Render lời giải dạng ngữ cảnh đơn giản."""
    formatted_content = format_simple_context_rich_text(explanation_json)
    cell.value = formatted_content
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    auto_size_cell(cell.parent, cell, plain_text_from_rich_text(formatted_content))

def render_shared_image_comprehension_explanation(cell, explanation_json: dict):
    """Render lời giải cho sheet 'TN PA đúng (img) (HL) (HSK1)'."""
    formatted_content = format_shared_image_comprehension_rich_text(explanation_json)
    cell.value = formatted_content
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    auto_size_cell(cell.parent, cell, plain_text_from_rich_text(formatted_content))

render_individual_img_explanation = render_simple_context_explanation
render_image_matching_explanation = render_simple_context_explanation
render_sentence_matching_explanation = render_simple_context_explanation

def render_hsk2_true_false_image_explanation(cell, explanation_json: dict):
    """
    Render lời giải cho dạng Đúng/Sai HSK2, bao gồm Phân tích, Phụ đề và Tạm dịch.
    """
    bold_font = InlineFont(b=True)
    italic_font = InlineFont(i=True)
    rich_text = CellRichText()

    # 1. Phân tích
    analysis = explanation_json.get('analysis_paragraph', '')
    rich_text.append(analysis + '\n\n')

    # 2. Phụ đề
    script_block = explanation_json.get('script_block', {})
    rich_text.append(TextBlock(bold_font, "Phụ đề:\n"))
    rich_text.append(script_block.get('chinese_text', '') + '\n')
    rich_text.append(script_block.get('pinyin', '') + '\n\n')

    # 3. Tạm dịch
    translation_block = explanation_json.get('translation_block', {})
    rich_text.append(TextBlock(bold_font, "Tạm dịch:\n"))
    rich_text.append(TextBlock(italic_font, translation_block.get('vietnamese_text', '')))

    # Gán giá trị và định dạng cho ô
    cell.value = rich_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Tự động điều chỉnh kích thước ô
    auto_size_cell(cell.parent, cell, plain_text_from_rich_text(rich_text))

def render_hsk2_dialogue_comprehension_explanation(cell, explanation_json: dict):
    """
    Render lời giải chung cho các dạng hội thoại HSK2, sử dụng mẫu thiết kế
    tạo plain text với markers rồi áp dụng regex formatting.
    """
    formatted_content = format_shared_image_comprehension_rich_text(explanation_json)
    cell.value = formatted_content
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    auto_size_cell(cell.parent, cell, plain_text_from_rich_text(formatted_content))

def render_hsk2_true_false_statement_explanation(cell, explanation_json: dict):
    """
    Render lời giải cho dạng Đúng/Sai từ nhận định HSK2.
    """
    bold_font = InlineFont(b=True)
    italic_font = InlineFont(i=True)
    rich_text = CellRichText()

    # 1. Phân tích
    analysis = explanation_json.get('analysis_paragraph', '')
    rich_text.append(analysis + '\n\n')

    # 2. Khối văn bản gốc
    original_texts = explanation_json.get('original_texts_block', {})
    rich_text.append(original_texts.get('script_chinese', '') + '\n')
    rich_text.append("★ " + original_texts.get('statement_chinese', '') + '\n\n')

    # 3. Khối tạm dịch
    translation = explanation_json.get('translation_block', {})
    rich_text.append(TextBlock(bold_font, "Tạm dịch:\n"))
    rich_text.append(TextBlock(italic_font,translation.get('script_vietnamese', '') + '\n'))
    rich_text.append(TextBlock(italic_font, "★ " + translation.get('statement_vietnamese', '')))

    # Gán giá trị và định dạng cho ô
    cell.value = rich_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Tự động điều chỉnh kích thước ô
    auto_size_cell(cell.parent, cell, plain_text_from_rich_text(rich_text))    

def render_hsk2_sentence_matching_inverted_explanation(cell, explanation_json: dict):
    """
    Render lời giải cho dạng ghép cặp câu trả lời HSK2.
    """
    bold_font = InlineFont(b=True)
    italic_font = InlineFont(i=True)
    rich_text = CellRichText()

    # 1. Phân tích
    analysis = explanation_json.get('analysis_paragraph', '')
    # Thêm Pinyin của câu thoại đúng (đáp án) vào sau lời giải thích
    matched_pair = explanation_json.get('matched_pair_block', {})
    statement_chinese = matched_pair.get('statement_chinese', '')
    statement_pinyin = matched_pair.get('statement_pinyin', '')
    analysis_with_pinyin = f"{analysis} (\"{statement_chinese}\" - {statement_pinyin})"
    rich_text.append(analysis_with_pinyin + '\n\n')

    # 2. Khối văn bản gốc (đáp án trước, câu hỏi sau)
    rich_text.append(matched_pair.get('statement_chinese', '') + '\n')
    rich_text.append(matched_pair.get('response_chinese', '') + '\n\n')

    # 3. Khối tạm dịch
    translation = explanation_json.get('translation_block', {})
    rich_text.append(TextBlock(bold_font, "Tạm dịch:\n"))
    # In nghiêng cả 2 dòng dịch
    rich_text.append(TextBlock(italic_font, translation.get('statement_vietnamese', '') + '\n'))
    rich_text.append(TextBlock(italic_font, translation.get('response_vietnamese', '')))

    # Gán giá trị và định dạng cho ô
    cell.value = rich_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Tự động điều chỉnh kích thước ô
    auto_size_cell(cell.parent, cell, plain_text_from_rich_text(rich_text))

def render_hsk3_true_false_listening_choice_explanation(cell, explanation_json: dict):
    """
    Render lời giải cho dạng Đúng/Sai nghe chọn HSK3.
    """
    bold_font = InlineFont(b=True)
    italic_font = InlineFont(i=True)
    rich_text = CellRichText()

    # 1. Phân tích
    analysis = explanation_json.get('analysis_paragraph', '')
    rich_text.append(analysis + '\n\n')

    # 2. Khối Phụ đề
    original_texts = explanation_json.get('original_texts_block', {})
    rich_text.append(TextBlock(bold_font, "Phụ đề:\n"))
    rich_text.append(original_texts.get('script_chinese', '') + '\n')
    # Loại bỏ dấu ★ nếu có sẵn trong dữ liệu và thêm lại để đảm bảo nhất quán
    statement = original_texts.get('statement_chinese', '').lstrip('★ ').strip()
    rich_text.append("★ " + statement + '\n\n')

    # 3. Khối Tạm dịch
    translation = explanation_json.get('translation_block', {})
    rich_text.append(TextBlock(bold_font, "Tạm dịch:\n"))
    # In nghiêng cả 2 dòng dịch
    rich_text.append(TextBlock(italic_font, translation.get('script_vietnamese', '') + '\n'))
    statement_vi = translation.get('statement_vietnamese', '').lstrip('★ ').strip()
    rich_text.append(TextBlock(italic_font, "★ " + statement_vi))

    # Gán giá trị và định dạng cho ô
    cell.value = rich_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Tự động điều chỉnh kích thước ô
    auto_size_cell(cell.parent, cell, plain_text_from_rich_text(rich_text))

def render_hsk3_reading_comprehension_no_pinyin_v3_explanation(cell, explanation_json: dict):
    """
    Render lời giải cho dạng đọc hiểu không pinyin V3 của HSK3 với rich text formatting.
    """
    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock, CellRichText
    from openpyxl.styles import Alignment
    
    bold_font = InlineFont(b=True)
    italic_font = InlineFont(i=True)
    rich_text = CellRichText()

    # 1. Phân tích
    analysis = explanation_json.get('analysis_paragraph', '')
    if analysis:
        rich_text.append(analysis)
        rich_text.append('\n\n')

    # 2. Khối Tạm dịch
    translation = explanation_json.get('translation_block', {})
    if translation:
        rich_text.append(TextBlock(bold_font, "Tạm dịch:\n"))
        script_chinese = translation.get('script_chinese', '')
        if script_chinese:
            rich_text.append(script_chinese + '\n')
        script_vietnamese = translation.get('script_vietnamese', '')
        if script_vietnamese:
            rich_text.append(TextBlock(italic_font, f"({script_vietnamese})"))
            rich_text.append('\n\n')

    # 3. Khối Câu hỏi và Lựa chọn
    query_options = explanation_json.get('query_and_options_block', {})
    
    # Câu hỏi
    query = query_options.get('query', {})
    if query:
        query_chinese = query.get('chinese_text', '').lstrip('★ ').strip()
        query_vietnamese = query.get('vietnamese_translation', '')
        if query_chinese:
            rich_text.append("★ " + query_chinese + '\n')
        if query_vietnamese:
            rich_text.append(TextBlock(italic_font, f"({query_vietnamese})"))
            rich_text.append('\n')

    # Các lựa chọn
    options = query_options.get('options', [])
    for opt in options:
        letter = opt.get('letter', '')
        chinese = opt.get('chinese_text', '')
        vietnamese = opt.get('vietnamese_translation', '')
        if letter and chinese:
            rich_text.append(f"{letter}. {chinese}\n")
        if vietnamese:
            rich_text.append(TextBlock(italic_font, f"({vietnamese})"))
            rich_text.append('\n')

    # Gán giá trị và định dạng cho ô
    cell.value = rich_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Tự động điều chỉnh kích thước ô
    auto_size_cell(cell.parent, cell, plain_text_from_rich_text(rich_text))

def build_hsk3_plain_text_with_markers(explanation_json: dict) -> str:
    """
    Tạo plain text với markers cho HSK3 reading comprehension (fallback method).
    """
    parts = []
    
    # 1. Phân tích
    analysis = explanation_json.get('analysis_paragraph', '')
    if analysis:
        parts.append(analysis)
        parts.append('\n\n')
    
    # 2. Tạm dịch
    translation = explanation_json.get('translation_block', {})
    if translation:
        parts.append('**Tạm dịch:**\n')
        script_chinese = translation.get('script_chinese', '')
        if script_chinese:
            parts.append(script_chinese + '\n')
        script_vietnamese = translation.get('script_vietnamese', '')
        if script_vietnamese:
            parts.append(f'<<i>>({script_vietnamese})<</i>>')
            parts.append('\n\n')
    
    # 3. Câu hỏi và lựa chọn
    query_options = explanation_json.get('query_and_options_block', {})
    query = query_options.get('query', {})
    if query:
        query_chinese = query.get('chinese_text', '').lstrip('★ ').strip()
        query_vietnamese = query.get('vietnamese_translation', '')
        if query_chinese:
            parts.append(f"★ {query_chinese}\n")
        if query_vietnamese:
            parts.append(f'<<i>>({query_vietnamese})<</i>>')
            parts.append('\n')
    
    options = query_options.get('options', [])
    for opt in options:
        letter = opt.get('letter', '')
        chinese = opt.get('chinese_text', '')
        vietnamese = opt.get('vietnamese_translation', '')
        if letter and chinese:
            parts.append(f"{letter}. {chinese}\n")
        if vietnamese:
            parts.append(f'<<i>>({vietnamese})<</i>>')
            parts.append('\n')
    
    return ''.join(parts)

def render_hsk3_reading_comprehension_no_pinyin_v3_explanation_with_markers(cell, explanation_json: dict):
    """
    Alternative renderer sử dụng marker system (tương tự các hàm khác trong file).
    """
    plain_text_with_markers = build_hsk3_plain_text_with_markers(explanation_json)
    rich_text = apply_regex_formatting(plain_text_with_markers)
    cell.value = rich_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    auto_size_cell(cell.parent, cell, plain_text_from_rich_text(rich_text))

def render_hsk3_sentence_reordering_explanation(cell, explanation_json: dict):
    """
    Render lời giải cho dạng sắp xếp thành phần câu HSK3.
    """
    bold_font = InlineFont(b=True)
    italic_font = InlineFont(i=True)
    rich_text = CellRichText()

    # 1. Phân tích ngữ pháp
    analysis = explanation_json.get('grammar_analysis', '')
    rich_text.append(analysis + '\n\n')

    # 2. Khối Tạm dịch
    correct_sentence = explanation_json.get('correct_sentence_block', {})
    rich_text.append(TextBlock(bold_font, "Tạm dịch:\n"))
    rich_text.append(TextBlock(italic_font, correct_sentence.get('vietnamese_translation', '')))

    # Gán giá trị và định dạng cho ô
    cell.value = rich_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Tự động điều chỉnh kích thước ô
    auto_size_cell(cell.parent, cell, plain_text_from_rich_text(rich_text))

def build_hsk4_text_simple(single_explanation: dict, shared_translation: dict, question_index: int, material_block: dict) -> str:
    """
    Tạo plain text đơn giản, đảm bảo markers được balance
    """
    
    # Khối 1: Phân tích
    analysis_block = single_explanation.get('analysis_paragraph', '')

    # Khối 2: Các lựa chọn - SAFE formatting
    options_lines = []
    options = single_explanation.get('options_list', [])
    for opt in options:
        letter = opt.get('letter', '')
        chinese = opt.get('chinese_text', '')
        translation = opt.get('translation', '').strip()
        if chinese and translation:
            options_lines.append(f"{letter}. {chinese} **({translation})")
    
    options_block = "\n".join(options_lines)

    # Khối 3: Phụ đề
    phude_block = ""
    script_chinese = material_block.get('script_chinese', '')
    if script_chinese:
        phude_block = f"**Phụ đề:\n{script_chinese}"

    # Khối 4: Tạm dịch  
    tamdich_block = ""
    translation_dict = shared_translation or {}
    script_vi = translation_dict.get('script_vietnamese', '')
    query_vi = translation_dict.get(f'query_vietnamese_{question_index + 1}', '')
    
    if script_vi or query_vi:
        translation_content = f"{script_vi}\n{query_vi}".strip()
        tamdich_block = f"**Tạm dịch:\n{translation_content}"

    # Ghép các khối
    final_parts = [analysis_block, options_block, phude_block, tamdich_block]
    return "\n\n".join(part for part in final_parts if part.strip())

def apply_hsk4_rich_text_formatting(text: str) -> CellRichText:
    """
    Quy tắc:
      - Bold: 'Phụ đề:' và 'Tạm dịch:'
      - Toàn bộ sau 'Tạm dịch:' -> italic
      - Bất kỳ (...) -> italic (bao gồm dấu ngoặc)
    Implementation: không append str trực tiếp, chỉ append TextBlock(InlineFont,...)
    """
    bold_font = InlineFont(b=True)
    italic_font = InlineFont(i=True)
    normal_font = InlineFont()

    rich = CellRichText()

    # Tokenize: tìm các keyword hoặc nhóm ngoặc đơn
    token_re = re.compile(r'(Phụ đề:|Tạm dịch:|\([^\)]*\))', flags=re.S)
    pos = 0
    in_tam_dich = False

    for m in token_re.finditer(text):
        before = text[pos:m.start()]
        if before:
            # append phần trước token theo trạng thái (italic nếu trong Tạm dịch)
            font = italic_font if in_tam_dich else normal_font
            rich.append(TextBlock(font, before))

        token = m.group(0)
        if token == "Phụ đề:":
            rich.append(TextBlock(bold_font, token))
        elif token == "Tạm dịch:":
            rich.append(TextBlock(bold_font, token))
            in_tam_dich = True
        else:
            # token là một nhóm ngoặc như "(...)" => luôn italic
            rich.append(TextBlock(italic_font, token))

        pos = m.end()

    # Phần còn lại sau token cuối
    tail = text[pos:]
    if tail:
        font = italic_font if in_tam_dich else normal_font
        rich.append(TextBlock(font, tail))

    return rich

def render_hsk4_listening_comprehension_explanation(worksheet, start_row: int, explanation_json: dict, task_data: dict):
    """
    Sử dụng build_hsk4_text_simple(...) để tạo plain text sạch (không có marker).
    Sau đó apply_hsk4_rich_text_formatting để gán CellRichText an toàn.
    """
    try:
        explanations = explanation_json.get('explanations', [])
        shared_translation = explanation_json.get('shared_translation', {})
        material_block = task_data.get('data', {})

        if len(explanations) != 2:
            print(f"⚠️ Cảnh báo: Nhận được {len(explanations)} lời giải thay vì 2.")
            # vẫn tiếp tục nếu muốn, hoặc return 0; giữ giống flow bạn dùng
            # return 0

        for i, single_explanation in enumerate(explanations):
            current_row = start_row + i
            cell = worksheet[f'I{current_row}']

            print(f"\n=== PROCESSING EXPLANATION {i+1} ===")

            # Lưu ý: dùng build_hsk4_text_simple (phiên bản bạn đang dùng)
            plain_text = build_hsk4_text_simple(single_explanation, shared_translation, i, material_block)

            # (Không replace marker ở đây — build_hsk4_text_simple phải trả về text sạch)
            print("Plain text preview:")
            print(plain_text[:200] + "..." if len(plain_text) > 200 else plain_text)

            # Apply rich text (rất quan trọng: luôn trả về TextBlock-only runs)
            rich_text = apply_hsk4_rich_text_formatting(plain_text)

            # Gán vào ô
            cell.value = rich_text
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        return len(explanations)

    except Exception as e:
        print(f"❌ Error in render_hsk4_listening_comprehension_explanation: {e}")
        # Fallback: ghi plain text (loại bỏ marker nếu có)
        for i, single_explanation in enumerate(explanations):
            current_row = start_row + i
            cell = worksheet[f'I{current_row}']
            plain_text = build_hsk4_text_simple(single_explanation, shared_translation, i, material_block)
            clean = plain_text.replace("**", "").replace("<<i>>", "").replace("<</i>>", "")
            cell.value = clean
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        return len(explanations)

# --- Cập nhật hàm builder text này ---
def build_hsk4_reordering_text_with_markers(explanation_json: dict, task_data: dict) -> str:
    """Tạo plain text với markers cho dạng sắp xếp câu HSK4. (Phiên bản đã sửa)"""
    parts = []
    
    # Lấy dữ liệu gốc để có nội dung câu
    components = task_data.get('data', {}).get('components', [])
    component_map = {comp['label']: comp['text'] for comp in components}
    correct_order = task_data.get('data', {}).get('correct_order', '')
    
    # 1. Khối giải thích logic
    logical_steps = explanation_json.get('logical_steps', [])
    step_lines = []
    for step in logical_steps:
        label = step.get('label', '')
        translation = step.get('translation', '')
        explanation = step.get('explanation', '')
        chinese_text = component_map.get(label, '')
        
        # Cấu trúc: - Câu [Label]: "[Tiếng Trung]" <<i>>([Bản dịch])<</i>> [Giải thích].
        step_lines.append(f"- Câu {label}: \"{chinese_text}\" <<i>>({translation})<</i>> {explanation}")
    
    if step_lines:
        parts.append("\n".join(step_lines))

    # 2. Khối đoạn văn hoàn chỉnh
    full_passage_parts = [component_map.get(letter, '') for letter in correct_order]
    full_passage_chinese = " ".join(filter(None, full_passage_parts))
    if full_passage_chinese:
        parts.append(f'\n\n{full_passage_chinese}')

    # 3. Khối tạm dịch
    full_translation = explanation_json.get('full_passage_translation', '')
    if full_translation:
        parts.append('\nTạm dịch:\n')
        parts.append(f'{full_translation}')

    # Dùng join để đảm bảo các phần được nối đúng cách
    return "".join(parts)

# --- Hàm renderer chính không cần thay đổi ---
def render_hsk4_sentence_sequence_reordering_explanation(cell, explanation_json: dict, task_data: dict):
    """
    Render lời giải cho dạng sắp xếp câu HSK4 bằng marker system.
    Cần task_data để lấy lại nội dung các câu gốc.
    """
    plain_text = build_hsk4_reordering_text_with_markers(explanation_json, task_data)
    rich_text = apply_rich_text_formatting(plain_text) # Tái sử dụng hàm formatter mạnh
    
    cell.value = rich_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    auto_size_cell(cell.parent, cell, plain_text_from_rich_text(rich_text))

# --- Thêm hàm builder text này ---
def build_hsk4_reading_passage_text_with_markers(single_explanation: dict, shared_translation: dict, question_index: int, material_block: dict) -> str:
    """Tạo plain text với markers cho MỘT lời giải trong cụm đọc hiểu ngắn HSK4."""
    
    # Lấy dữ liệu cần thiết
    analysis = single_explanation.get('analysis_paragraph', '')
    options = single_explanation.get('options_list', [])
    script_chinese = material_block.get('script_chinese', '')
    questions_data = material_block.get('questions', [{}, {}])
    current_question = questions_data[question_index]
    query_chinese = current_question.get('query_chinese', '')

    passage_vi = shared_translation.get('script_vietnamese', '')
    query_vi = shared_translation.get(f'query_vietnamese_{question_index + 1}', '')

    # --- Xây dựng các khối ---
    
    # Khối 1: Phân tích
    analysis_block = analysis

    # Khối 2: Tạm dịch học liệu
    tamdich_passage_block = ""
    if script_chinese:
        tamdich_passage_block = f"**Tạm dịch:**\n{script_chinese}\n<<i>>({passage_vi})<</i>>"

    # Khối 3: Câu hỏi và các lựa chọn
    question_block_lines = []
    if query_chinese:
        question_block_lines.append(f"{query_chinese}")
        question_block_lines.append(f"<<i>>({query_vi})<</i>>")
    
    if options:
        for opt in options:
            letter, chinese, translation = opt.get('letter', ''), opt.get('chinese_text', ''), opt.get('translation', '')
            question_block_lines.append(f"{letter}. {chinese}\n<<i>>({translation})<</i>>")
    
    question_block = "\n".join(question_block_lines)

    # --- Ghép các khối ---
    final_parts = [
        analysis_block,
        tamdich_passage_block,
        question_block
    ]
    
    return "\n\n".join(part for part in final_parts if part)

# --- Thêm hàm renderer chính này ---
def render_hsk4_reading_passage_explanation(worksheet, start_row: int, explanation_json: dict, task_data: dict):
    """Render lời giải cho một cụm đọc hiểu ngắn HSK4, ghi ra 2 dòng liên tiếp."""
    explanations = explanation_json.get('explanations', [])
    shared_translation = explanation_json.get('shared_translation', {})
    material_block = task_data.get('data', {})

    if len(explanations) != 2:
        print(f"   ⚠️ Cảnh báo: Nhận được {len(explanations)} lời giải thay vì 2. Bỏ qua.")
        return 0

    for i, single_explanation in enumerate(explanations):
        current_row = start_row + i
        cell = worksheet[f'I{current_row}']
        
        plain_text = build_hsk4_reading_passage_text_with_markers(single_explanation, shared_translation, i, material_block)
        rich_text = apply_regex_formatting(plain_text)
        
        cell.value = rich_text
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        auto_size_cell(worksheet, cell, plain_text_from_rich_text(rich_text))

    return 2

# --- Thêm hàm builder text này ---
def build_hsk4_image_word_sentence_text_with_markers(explanation_json: dict, task_data: dict) -> str:
    """Tạo plain text với markers cho lời giải dạng Đặt câu với ảnh và từ HSK4."""
    
    # Lấy dữ liệu từ AI
    translation = explanation_json.get('sample_sentence_translation', '')
    explanation = explanation_json.get('explanation_paragraph', '')

    # Lấy dữ liệu gốc từ task_data
    sample_sentence = task_data.get('data', {}).get('sample_sentence', '')

    # --- Xây dựng các khối ---
    
    # Khối 1: Đáp án
    dapan_block = f"**Đáp án:** {sample_sentence}"

    # Khối 2: Tạm dịch
    tamdich_block = f"**Tạm dịch:** <<i>>{translation}<</i>>"
    
    # Khối 3: Giải thích
    giaithich_block = f"**Giải thích:** {explanation}"

    # --- Ghép các khối ---
    final_parts = [
        dapan_block,
        tamdich_block,
        giaithich_block
    ]
    
    return "\n\n".join(part for part in final_parts if part)

# --- Thêm hàm renderer chính này ---
def render_hsk4_image_word_sentence_creation_explanation(cell, explanation_json: dict, task_data: dict):
    """
    Render lời giải cho dạng Đặt câu với ảnh và từ HSK4 bằng cách xây dựng RichText trực tiếp.
    - In đậm: "Đáp án:", "Tạm dịch:", "Giải thích:"
    - In nghiêng: Nội dung sau "Tạm dịch:"
    """
    # 1. Định nghĩa các kiểu font
    bold_font = InlineFont(b=True)
    italic_font = InlineFont(i=True)

    # 2. Lấy dữ liệu cần thiết
    # Dữ liệu từ AI
    translation = explanation_json.get('sample_sentence_translation', '')
    explanation = explanation_json.get('explanation_paragraph', '')
    # Dữ liệu gốc từ task_data
    sample_sentence = task_data.get('data', {}).get('sample_sentence', '')
    
    # 3. Xây dựng đối tượng CellRichText từng phần
    rich_text = CellRichText()

    # Phần 1: Đáp án
    if sample_sentence:
        rich_text.append(TextBlock(bold_font, 'Đáp án: '))
        rich_text.append(f'{sample_sentence}\n\n')

    # Phần 2: Tạm dịch
    if translation:
        rich_text.append(TextBlock(bold_font, 'Tạm dịch: '))
        rich_text.append(TextBlock(italic_font, f'{translation}\n\n'))

    # Phần 3: Giải thích
    if explanation:
        rich_text.append(TextBlock(bold_font, 'Giải thích: '))
        rich_text.append(explanation)

    # 4. Gán giá trị và định dạng cho ô
    cell.value = rich_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Lấy văn bản thô để tính toán kích thước
    plain_text_for_sizing = plain_text_from_rich_text(rich_text)
    auto_size_cell(cell.parent, cell, plain_text_for_sizing)

# --- Thêm hàm builder text này ---
def build_hsk5_passage_cloze_text_with_markers(single_explanation: dict, shared_translation: dict, full_passage_chinese: str) -> str:
    """Tạo plain text với markers cho MỘT lời giải trong cụm điền từ HSK5."""

    # Lấy dữ liệu cần thiết
    analysis = single_explanation.get('analysis_paragraph', '')
    options = single_explanation.get('options_list', [])
    full_translation = shared_translation.get('full_passage_vietnamese', '')

    # --- Xây dựng các khối ---
    
    # Khối 1: Phân tích
    analysis_block = analysis

    # Khối 2: Tạm dịch (chung cho cả 4 câu)
    tamdich_block = ""
    if full_passage_chinese and full_translation:
        tamdich_block = (
            f"**Tạm dịch:**\n{full_passage_chinese}\n"
            f"<<i>>({full_translation})<</i>>"
        )

    # Khối 3: Các lựa chọn
    options_block_lines = []
    if options:
        for opt in options:
            letter, chinese, translation = opt.get('letter', ''), opt.get('chinese_text', ''), opt.get('translation', '')
            options_block_lines.append(f"{letter}. {chinese} \n<<i>>({translation})<</i>>")
    options_block = "\n".join(options_block_lines)

    # --- Ghép các khối ---
    final_parts = [analysis_block, tamdich_block, options_block]
    return "\n\n".join(part for part in final_parts if part)

def apply_regex_formatting_with_highlight(plain_text: str) -> CellRichText:
    """
    Parse các marker và build CellRichText.
    Xử lý:
    - <<bi>>...<</bi>> (In đậm + nghiêng)
    - <<i>>...<</i>> (In nghiêng)
    - **...** (In đậm)
    - __...__ (Màu xanh lá cây)
    """
    bold_font        = InlineFont(b=True)
    italic_font      = InlineFont(i=True)
    bold_italic_font = InlineFont(b=True, i=True)
    green_font       = InlineFont(color="008000") # Màu xanh lá cây đậm

    rich = CellRichText()

    # Regex kết hợp tất cả các loại marker
    # (?s) để cho phép '.' khớp với ký tự xuống dòng
    # (?:...) là non-capturing group
    token_re = re.compile(r'(?s)(?:(<<bi>>.*?<<\/bi>>)|(<<i>>.*?<<\/i>>)|(\*\*[^*]+\*\*)|(__.*?__))')

    pos = 0
    for m in token_re.finditer(plain_text):
        # 1. Thêm phần text thường nằm trước token
        if m.start() > pos:
            rich.append(plain_text[pos:m.start()])

        # 2. Xử lý token đã tìm thấy
        # Lấy token từ group khớp đầu tiên không phải None
        tok = next(g for g in m.groups() if g is not None)
        
        if tok.startswith('<<bi>>'):
            content = tok[6:-7] # Bỏ <<bi>> và <</bi>>
            rich.append(TextBlock(bold_italic_font, content))
        elif tok.startswith('<<i>>'):
            content = tok[5:-6] # Bỏ <<i>> và <</i>>
            rich.append(TextBlock(italic_font, content))
        elif tok.startswith('**'):
            content = tok[2:-2] # Bỏ **
            rich.append(TextBlock(bold_font, content))
        elif tok.startswith('__'):
            content = tok[2:-2] # Bỏ __
            rich.append(TextBlock(green_font, content))

        pos = m.end()

    # 3. Thêm phần text còn lại sau token cuối cùng
    if pos < len(plain_text):
        rich.append(plain_text[pos:])

    return rich

# --- Thêm hàm renderer chính này ---
def render_hsk5_passage_cloze_explanation(worksheet, start_row: int, explanation_json: dict, task_data: dict):
    """Render lời giải cho một cụm điền từ HSK5, ghi ra 4 dòng liên tiếp."""
    
    # Dữ liệu từ AI
    explanations = explanation_json.get('explanations', [])
    shared_translation = explanation_json.get('shared_translation', {})
    
    # Dữ liệu gốc
    passage_block = task_data.get('data', {})
    passage_with_blanks = passage_block.get('shared_material', '')
    questions = passage_block.get('questions', [])

    if len(explanations) != 4 or len(questions) != 4:
        print(f"   ⚠️ Cảnh báo: Dữ liệu HSK5 Cloze không đủ. Bỏ qua.")
        return 0

    # Tái tạo lại đoạn văn hoàn chỉnh một lần
    correct_words = [q.get('answer_options', [])[q.get('correct_answer', 1)-1] for q in questions]
    full_passage_chinese = passage_with_blanks
    for word in correct_words:
        full_passage_chinese = full_passage_chinese.replace('________', f'__{word}__', 1)

    # Lặp qua 4 lời giải và ghi vào 4 dòng
    for i, single_explanation in enumerate(explanations):
        current_row = start_row + i
        cell = worksheet[f'I{current_row}']
        
        plain_text = build_hsk5_passage_cloze_text_with_markers(single_explanation, shared_translation, full_passage_chinese)
        rich_text = apply_regex_formatting_with_highlight(plain_text)
        
        cell.value = rich_text
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        auto_size_cell(worksheet, cell, plain_text_from_rich_text(rich_text))

    return 4

# --- Thêm hàm builder text này ---
def build_hsk5_main_idea_text_with_markers(explanation_json: dict, task_data: dict) -> str:
    """Tạo plain text với markers cho dạng chọn chủ đề đoạn văn HSK5."""

    # Lấy dữ liệu từ AI
    analysis = explanation_json.get('analysis_paragraph', '')
    passage_translation = explanation_json.get('passage_translation', '')
    options = explanation_json.get('options_list_with_translation', [])
    
    # Lấy dữ liệu gốc
    passage_chinese = task_data.get('data', {}).get('passage_chinese', '')

    # --- Xây dựng các khối ---
    
    # Khối 1: Phân tích
    analysis_block = analysis

    # Khối 2: Tạm dịch
    tamdich_block = ""
    if passage_chinese and passage_translation:
        tamdich_block = (
            f"**Tạm dịch:**\n{passage_chinese}\n"
            f"<<i>>({passage_translation})<</i>>"
        )

    # Khối 3: Các lựa chọn
    options_block_lines = []
    if options:
        for opt in options:
            letter, chinese, translation = opt.get('letter', ''), opt.get('chinese_text', ''), opt.get('translation', '')
            options_block_lines.append(f"{letter}. {chinese} \n<<i>>({translation})<</i>>")
    options_block = "\n".join(options_block_lines)

    # --- Ghép các khối ---
    final_parts = [analysis_block, tamdich_block, options_block]
    return "\n\n".join(part for part in final_parts if part)

# --- Thêm hàm renderer chính này ---
def render_hsk5_main_idea_comprehension_explanation(cell, explanation_json: dict, task_data: dict):
    """
    Render lời giải cho dạng chọn chủ đề đoạn văn HSK5.
    """
    plain_text = build_hsk5_main_idea_text_with_markers(explanation_json, task_data)
    rich_text = apply_regex_formatting(plain_text) # Tái sử dụng hàm formatter mạnh
    
    cell.value = rich_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    auto_size_cell(cell.parent, cell, plain_text_from_rich_text(rich_text))

# --- Thêm hàm builder text này ---
def build_hsk5_long_passage_text_with_markers(single_explanation: dict, shared_translation: dict, passage_chinese: str) -> str:
    """Tạo plain text với markers cho MỘT lời giải trong cụm đọc hiểu dài HSK5."""

    # Lấy dữ liệu cần thiết
    analysis = single_explanation.get('analysis_paragraph', '')
    options = single_explanation.get('options_list', [])
    passage_translation = shared_translation.get('passage_vietnamese', '')

    # --- Xây dựng các khối ---
    
    # Khối 1: Phân tích
    analysis_block = analysis

    # Khối 2: Tạm dịch (chung cho cả 4 câu)
    tamdich_block = ""
    if passage_chinese and passage_translation:
        tamdich_block = (
            f"**Tạm dịch:**\n{passage_chinese}\n"
            f"<<i>>({passage_translation})<</i>>"
        )

    # Khối 3: Các lựa chọn
    options_block_lines = []
    if options:
        for opt in options:
            letter, chinese, translation = opt.get('letter', ''), opt.get('chinese_text', ''), opt.get('translation', '')
            options_block_lines.append(f"{letter}. {chinese} \n<<i>>({translation})<</i>>")
    options_block = "\n".join(options_block_lines)

    # --- Ghép các khối ---
    final_parts = [analysis_block, tamdich_block, options_block]
    return "\n\n".join(part for part in final_parts if part)

# --- Thêm hàm renderer chính này ---
def render_hsk5_long_passage_comprehension_explanation(worksheet, start_row: int, explanation_json: dict, task_data: dict):
    """Render lời giải cho một cụm đọc hiểu dài HSK5, ghi ra 4 dòng liên tiếp."""
    
    explanations = explanation_json.get('explanations', [])
    shared_translation = explanation_json.get('shared_translation', {})
    passage_block = task_data.get('data', {})
    passage_chinese = passage_block.get('shared_passage', '')

    if len(explanations) != 4:
        print(f"   ⚠️ Cảnh báo: Dữ liệu HSK5 Long Passage không đủ. Bỏ qua.")
        return 0

    # Lặp qua 4 lời giải và ghi vào 4 dòng
    for i, single_explanation in enumerate(explanations):
        current_row = start_row + i
        cell = worksheet[f'I{current_row}']
        
        plain_text = build_hsk5_long_passage_text_with_markers(single_explanation, shared_translation, passage_chinese)
        rich_text = apply_regex_formatting(plain_text)
        
        cell.value = rich_text
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        auto_size_cell(worksheet, cell, plain_text_from_rich_text(rich_text))

    return 4

def apply_formatting_with_highlight_and_headings(text: str) -> CellRichText:
    """
    Áp dụng định dạng rich text:
    - Bôi đậm các tiêu đề "Tạm dịch:", "Phụ đề:", v.v.
    - In nghiêng text sau "Tạm dịch:".
    - Tô màu xanh lá cây cho text trong __...__ (và xóa __).
    """
    # 1. Định nghĩa các kiểu font
    bold_font = InlineFont(b=True)
    italic_font = InlineFont(i=True)
    green_font = InlineFont(color="008000")

    rich_text = CellRichText()
    
    # 2. Định nghĩa các tiêu đề cần bôi đậm
    bold_headings = ["Tạm dịch:", "Phụ đề:"] # Có thể mở rộng danh sách này
    
    # 3. Tạo pattern regex - SỬA LẠI ĐỂ TRÁNH NESTED GROUPS
    # Không cần group cho headings_pattern nữa
    headings_pattern = '|'.join(re.escape(kw) for kw in bold_headings)
    highlight_pattern = r'__.*?__'
    # Kết hợp pattern
    split_pattern = f'({headings_pattern}|{highlight_pattern})'

    # 4. Tách chuỗi theo cả tiêu đề và highlight
    parts = re.split(split_pattern, text)
    
    in_tam_dich_section = False # Cờ để theo dõi trạng thái in nghiêng
    
    for part in filter(None, parts): # Lọc ra các chuỗi rỗng
        is_heading = part in bold_headings
        is_highlight = part.startswith('__') and part.endswith('__')

        if is_heading:
            rich_text.append(TextBlock(bold_font, part))
            if part == "Tạm dịch:":
                in_tam_dich_section = True
            else:
                # Reset cờ nếu là tiêu đề khác (ví dụ "Phụ đề:")
                in_tam_dich_section = False
        elif is_highlight:
            content = part[2:-2] # Bỏ dấu __
            rich_text.append(TextBlock(green_font, content))
        else:
            # Đây là văn bản thường
            if in_tam_dich_section:
                rich_text.append(TextBlock(italic_font, part))
            else:
                rich_text.append(part)
    
    return rich_text

def build_hsk5_writing_from_keywords_text_with_markers(question_data: dict) -> str:
    """Tạo plain text với markers cho lời giải dạng Viết dựa vào từ HSK5."""
    
    # Lấy dữ liệu cần thiết từ question_data (vì lời giải đã có sẵn)
    paragraph_marked = question_data.get('sample_paragraph_marked', '')
    translation = question_data.get('translation', '')

    # --- Xây dựng các khối ---
    
    # Khối 1: Đoạn văn mẫu (đã có sẵn marker __...__)
    paragraph_block = paragraph_marked

    # Khối 2: Tạm dịch - KIỂM TRA XEM ĐÃ CÓ "Tạm dịch:" CHƯA
    tamdich_block = ""
    if translation:
        # Kiểm tra nếu translation đã có "Tạm dịch:" thì không thêm nữa
        if translation.strip().startswith("Tạm dịch:"):
            tamdich_block = translation
        else:
            tamdich_block = f"Tạm dịch:\n{translation}"
        
    # --- Ghép các khối ---
    final_parts = [paragraph_block, tamdich_block]
    return "\n\n".join(part for part in final_parts if part)

def render_hsk5_writing_from_keywords(cell, question_data: dict):
    """
    Render lời giải cho dạng Viết dựa vào từ HSK5 bằng marker system.
    """
    # 1. Tạo plain text với các marker
    plain_text = build_hsk5_writing_from_keywords_text_with_markers(question_data)
    
    # 2. Áp dụng định dạng bằng formatter đa năng
    rich_text = apply_formatting_with_highlight_and_headings(plain_text)
    
    # 3. Gán giá trị và định dạng cho ô
    cell.value = rich_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    auto_size_cell(cell.parent, cell, plain_text_from_rich_text(rich_text))

def build_hsk5_writing_from_image_text_with_markers(question_data: dict) -> str:
    """Tạo plain text với markers cho lời giải dạng Viết dựa vào ảnh HSK5."""

    # Lấy dữ liệu cần thiết từ question_data
    sample_paragraph = question_data.get('sample_paragraph', '')
    translation = question_data.get('translation', '')

    # --- Xây dựng các khối ---

    # Khối 1: Đoạn văn mẫu
    paragraph_block = sample_paragraph
    
    # Khối 2: Tạm dịch
    tamdich_block = ""
    if translation:
        tamdich_block = f"Tạm dịch:\n {translation}"
        
    # --- Ghép các khối ---
    final_parts = [paragraph_block, tamdich_block]
    return "\n\n".join(part for part in final_parts if part)    

def render_hsk5_writing_from_image(cell, question_data: dict):
    """
    Render lời giải cho dạng Viết dựa vào ảnh HSK5 bằng marker system.
    """
    # 1. Tạo plain text với các marker
    plain_text = build_hsk5_writing_from_image_text_with_markers(question_data)

    # 2. Áp dụng định dạng
    # Dùng hàm apply_regex_formatting là đủ vì không có highlight xanh
    rich_text = apply_ds_vocab_rich_text_formatting(plain_text)
    
    # 3. Gán giá trị và định dạng cho ô
    cell.value = rich_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    auto_size_cell(cell.parent, cell, plain_text_from_rich_text(rich_text))


def render_topik_word_matching_explanation(worksheet, current_row, explanation_data, original_task=None):
    """
    Render lời giải nối từ. 
    Lưu ý: Vì dạng này không gọi AI tạo lời giải, dữ liệu nằm trong original_task['data']
    """
    # Lấy dữ liệu từ task gốc (do config HSK5 Writing logic)
    data_source = explanation_data if explanation_data else (original_task['data'] if original_task else None)
    
    if not data_source: return 0

    pairs = data_source.get('pairs', [])
    
    # Format: 한국: Hàn Quốc (xuống dòng)
    explanation_lines = []
    for p in pairs:
        explanation_lines.append(f"{p['korean']}: {p['vietnamese']}")
    
    final_text = "\n".join(explanation_lines)
    
    # Ghi vào cột I
    worksheet[f'I{current_row}'] = final_text
    
    return 1 # Trả về số dòng đã ghi (1 dòng Excel)

def render_topik_image_matching_explanation(worksheet, current_row, explanation_data, original_task=None):
    """
    Render lời giải nối ảnh (Hàn - Việt).
    """
    data_source = explanation_data if explanation_data else (original_task['data'] if original_task else None)
    if not data_source: return 0

    pairs = data_source.get('pairs', [])
    
    # Format: 일본: Nhật Bản
    explanation_lines = []
    for p in pairs:
        explanation_lines.append(f"{p['korean']}: {p['vietnamese']}")
    
    final_text = "\n".join(explanation_lines)
    
    # Ghi vào cột I
    worksheet[f'I{current_row}'] = final_text
    
    return 1

def render_topik_image_selection_explanation(worksheet, current_row, explanation_data, original_task=None):
    """
    Render lời giải TN chọn ảnh.
    Hiển thị danh sách từ vựng của 4 phương án.
    """
    # Lấy data từ original_task vì không gọi AI
    data_source = explanation_data if explanation_data else (original_task['data'] if original_task else None)
    if not data_source: return 0

    options = data_source.get('options', [])
    
    # Format: 
    # 극장: Nhà hát
    # 은행: Ngân hàng
    explanation_lines = []
    for opt in options:
        line = f"{opt.get('korean', '')}: {opt.get('vietnamese', '')}"
        explanation_lines.append(line)
    
    final_text = "\n".join(explanation_lines)
    
    # Ghi vào cột I
    worksheet[f'I{current_row}'] = final_text
    
    return 1

def render_topik_image_to_word_explanation(worksheet, current_row, explanation_data, original_task=None):
    """
    Render lời giải: Liệt kê nghĩa của 4 phương án.
    """
    data_source = explanation_data if explanation_data else (original_task['data'] if original_task else None)
    if not data_source: return 0

    options = data_source.get('options', [])
    
    # Format: 
    # 운동장: Sân vận động
    explanation_lines = []
    for opt in options:
        line = f"{opt.get('korean', '')}: {opt.get('vietnamese', '')}"
        explanation_lines.append(line)
    
    final_text = "\n".join(explanation_lines)
    worksheet[f'I{current_row}'] = final_text
    
    return 1

def render_topik_vocab_selection_explanation(worksheet, current_row, explanation_data, original_task=None):
    """
    Render lời giải: display_text : meaning
    Ví dụ Type 1: 그저께: Hôm kia
    Ví dụ Type 2: Thứ hai: 월요일
    """
    data_source = explanation_data if explanation_data else (original_task['data'] if original_task else None)
    if not data_source: return 0

    options = data_source.get('options', [])
    
    explanation_lines = []
    for opt in options:
        # Cấu trúc: [Hiển thị ở cột G]: [Nghĩa giải thích]
        line = f"{opt.get('display_text', '')}: {opt.get('meaning', '')}"
        explanation_lines.append(line)
    
    final_text = "\n".join(explanation_lines)
    worksheet[f'I{current_row}'] = final_text
    
    return 1

def render_topik_reading_shared_explanation(worksheet, current_row, explanation_data, original_task=None):
    """
    Render lời giải Đọc hiểu chung.
    Format:
    Câu hỏi: ... (...)
    + Opt 1: ...
    => Dựa vào ...
    
    Tạm dịch:
    (Bản dịch bài đọc)
    """
    # Vì ta đã lưu dữ liệu vào original_task ở bước populate
    # explanation_data có thể là None do ta skip API
    task_data = original_task['data'] if original_task else None
    if not task_data: return 0

    # Lấy các thông tin đã lưu
    q_text = task_data.get('question_text', '')
    q_mean = task_data.get('question_meaning', '')
    evidence = task_data.get('reasoning_evidence', '')
    evidence_mean = task_data.get('reasoning_vietnamese', '') # Nếu prompt có sinh ra
    passage_trans = task_data.get('shared_passage_translation', '')
    options = task_data.get('options', [])

    # 1. Phần Câu hỏi và Đáp án
    lines = []
    lines.append(f"Câu hỏi: {q_text} ({q_mean})")
    
    for opt in options:
        lines.append(f"+ {opt['korean']}: {opt['vietnamese']}")
    
    # 2. Phần Dẫn chứng
    lines.append(f"=> Dựa vào {evidence} ({evidence_mean})")
    lines.append("") # Dòng trống
    
    # 3. Phần Tạm dịch (Học liệu chung)
    lines.append("Tạm dịch:")
    lines.append(passage_trans)

    final_text = "\n".join(lines)
    
    # Ghi vào cột I
    worksheet[f'I{current_row}'] = final_text
    worksheet[f'I{current_row}'].alignment = Alignment(wrap_text=True)

    return 1

def render_topik_listening_vocab_explanation(worksheet, current_row, explanation_data, original_task=None):
    """
    Render lời giải: Tiếng Việt: Tiếng Hàn /Note/
    Ví dụ: Tháng tư: 4월 /사월/
    """
    data_source = explanation_data if explanation_data else (original_task['data'] if original_task else None)
    if not data_source: return 0

    options = data_source.get('options', [])
    
    explanation_lines = []
    for opt in options:
        vn = opt.get('vietnamese', '')
        kr = opt.get('korean_equiv', '')
        note = opt.get('note', '') # Có thể rỗng
        
        # Xây dựng chuỗi: "Thứ tư: 수요일"
        line = f"{vn}: {kr}"
        
        # Nếu có note thì thêm vào: "Tháng hai: 2월 /이월/"
        if note:
            line += f" {note}"
            
        explanation_lines.append(line)
    
    final_text = "\n".join(explanation_lines)
    worksheet[f'I{current_row}'] = final_text
    
    return 1

def render_topik_listening_fill_explanation(worksheet, current_row, explanation_data, original_task=None):
    """
    Render lời giải:
    Word: Meaning
    Full Sentence
    Tạm dịch: ...
    """
    data_source = explanation_data if explanation_data else (original_task['data'] if original_task else None)
    if not data_source: return 0

    target = data_source.get('target_word', '')
    meaning = data_source.get('target_meaning', '')
    full_sentence = data_source.get('full_sentence', '')
    trans = data_source.get('sentence_translation', '')

    lines = []
    # Dòng 1: Từ vựng
    lines.append(f"{target}: {meaning}")
    # Dòng 2: Câu gốc
    lines.append(full_sentence)
    # Dòng 3: Dịch
    lines.append(f"Tạm dịch: {trans}")

    final_text = "\n".join(lines)
    worksheet[f'I{current_row}'] = final_text
    
    return 1

def render_topik_listening_reorder_explanation(worksheet, current_row, explanation_data, original_task=None):
    """
    Render lời giải:
    Từ 1: Nghĩa 1
    Từ 2: Nghĩa 2
    ...
    Chuỗi từ đúng hoàn chỉnh
    """
    data_source = explanation_data if explanation_data else (original_task['data'] if original_task else None)
    if not data_source: return 0

    # Lấy danh sách đúng thứ tự
    items = data_source.get('items', []) # hoặc 'sorted_items'
    
    lines = []
    correct_koreans = []

    # Liệt kê nghĩa
    for item in items:
        lines.append(f"{item['korean']}: {item['vietnamese']}")
        correct_koreans.append(item['korean'])
    
    # Dòng cuối: Chuỗi hoàn chỉnh
    lines.append(", ".join(correct_koreans))

    final_text = "\n".join(lines)
    worksheet[f'I{current_row}'] = final_text
    
    return 1

def render_topik_listening_shared_explanation(worksheet, current_row, explanation_data, original_task=None):
    """
    Render lời giải Nghe hiểu O/X.
    """
    # Lấy data từ original_task (do Phase 1 tạo ra)
    task_data = original_task['data'] if original_task else None
    if not task_data: return 0

    # 1. Lấy dữ liệu (với giá trị mặc định để tránh lỗi None)
    stmt_kr = task_data.get('statement_korean', '')
    stmt_vn = task_data.get('statement_vietnamese', '')
    
    # Logic Đúng/Sai từ trường is_correct
    is_correct = task_data.get('is_correct', False) 
    
    # Lấy giải thích và dẫn chứng
    expl = task_data.get('explanation', '')
    evidence = task_data.get('evidence_korean', '')
    
    # Lấy hội thoại (đã được gán ở bước flatten)
    dialogue = task_data.get('full_dialogue_data', [])

    lines = []
    
    # --- PHẦN 1: GIẢI THÍCH ---
    # In câu nhận định
    lines.append(f"{stmt_kr} ({stmt_vn})")
    
    # In lý do Đúng/Sai
    # Nếu is_correct = True (Đáp án O) -> Hiển thị "Đúng vì..."
    # Nếu is_correct = False (Đáp án X) -> Hiển thị "Sai vì..."
    status_text = "Đúng" if is_correct else "Sai"
    
    # Chỉ in dẫn chứng nếu có dữ liệu
    if evidence or expl:
        lines.append(f"{status_text} vì {evidence} ({expl})")
    else:
        lines.append(f"{status_text}.")
        
    lines.append("") # Dòng trống ngăn cách

    # --- PHẦN 2: PHỤ ĐỀ (HÀN) ---
    lines.append("Phụ đề")
    if dialogue:
        for d in dialogue:
            speaker = d.get('speaker', '')
            text = d.get('korean', '')
            lines.append(f"{speaker}: {text}")
    else:
        lines.append("(Không có dữ liệu phụ đề)")
    lines.append("")

    # --- PHẦN 3: TẠM DỊCH (VIỆT) ---
    lines.append("Tạm dịch")
    if dialogue:
        for d in dialogue:
            # Chuyển đổi tên nhân vật sang tiếng Việt cho thân thiện
            speaker_kr = d.get('speaker', '')
            speaker_vn = "Nam" if "남" in speaker_kr else ("Nữ" if "여" in speaker_kr else speaker_kr)
            
            text_vn = d.get('vietnamese', '')
            lines.append(f"{speaker_vn}: {text_vn}")
    else:
        lines.append("(Không có dữ liệu dịch)")

    final_text = "\n".join(lines)
    
    # Ghi vào cột I
    worksheet[f'I{current_row}'] = final_text
    from openpyxl.styles import Alignment
    worksheet[f'I{current_row}'].alignment = Alignment(wrap_text=True)

    return 1