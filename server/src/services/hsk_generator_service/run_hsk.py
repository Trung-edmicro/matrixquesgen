import os
import json
import shutil
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any
import openpyxl
from .utils.call_vertex_ai import ai_client, get_resource_path
from .utils.matrix_manager import MatrixManager
from .config.hsk_question_configs import get_prompt_config
from .config.hsk_explanation_configs import get_explanation_config
from .formatters.explanation_excel_formatter import render_ds_vocab_explanation

class HSKPipeline:
    def __init__(self, hsk_level: str):
        self.hsk_level = hsk_level.lower()
        
        # Thiết lập đường dẫn
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.output_dir = get_resource_path(f"output/{self.hsk_level}_{timestamp}")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.excel_output_path = self.output_dir / f"{self.hsk_level}_result.xlsx"
        self.json_data_path = self.output_dir / "generated_data.json"
        
        # 1. Khởi tạo file Excel từ Template
        self._prepare_excel()
        
    def _prepare_excel(self):
        """Copy file excel mẫu từ resources/sheet sang output"""
        template_path = get_resource_path(f"resources/sheet/{self.hsk_level}.xlsx")
        if not template_path.exists():
            raise FileNotFoundError(f"❌ Không tìm thấy template: {template_path}")
        shutil.copy(template_path, self.excel_output_path)
        print(f"✅ Đã chuẩn bị file Excel tại: {self.excel_output_path}")

    def _load_resource(self, sub_path: str):
        """Helper để đọc file prompt/schema"""
        path = get_resource_path(sub_path)
        if path.suffix == '.json':
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        return path.read_text(encoding='utf-8')

    def run_question_generation(self, matrix_manager):
        """Bước 2: Tạo câu hỏi và điền vào Excel dựa trên cấu trúc ma trận"""
        print(f"\n--- [BƯỚC 2] TẠO CÂU HỎI ---")
        configs = get_prompt_config(self.hsk_level)
        workbook = openpyxl.load_workbook(self.excel_output_path)
        
        all_generated_data = {}

        for prompt_id, config in configs.items():
            matrix_name = config.get("matrix_name")
            if not matrix_name:
                print(f"   ⏩ Bỏ qua {prompt_id} (Không có cấu hình matrix_name)")
                continue

            # 1. Lấy dữ liệu động từ matrix_manager
            dynamic_data = matrix_manager.get_dynamic_data_for_form(self.hsk_level, matrix_name)
            if not dynamic_data:
                print(f"   ⏩ Bỏ qua {prompt_id} (Không tìm thấy '{matrix_name}' trong file matrix)")
                continue

            print(f"   🚀 Đang chạy dạng bài: {matrix_name} ({prompt_id})")
            
            # 2. Đọc file text template và format
            p_text_template = self._load_resource(f"resources/prompts/create/{self.hsk_level}/{prompt_id}.txt")
            final_prompt = p_text_template.format(
                number=dynamic_data["number"],
                matrix_name=dynamic_data["matrix_name"],
                rules_text=dynamic_data["rules_text"],
                vocab_md=dynamic_data["vocab_md"]
            )
            
            # 3. Đọc Schema
            p_schema = self._load_resource(f"resources/schemas/create/{self.hsk_level}/{prompt_id}.json")
            
            # 4. Gọi AI
            data = ai_client.generate_content(final_prompt, p_schema)
            all_generated_data[prompt_id] = data
            
            # 5. Ghi dữ liệu vào Excel (Duyệt qua các processors trong config)
            for json_key, sheet_name, populate_func in config["processors"]:
                if sheet_name in workbook.sheetnames:
                    data_part = data.get(json_key)
                    if data_part:
                        populate_func(workbook[sheet_name], data_part)
                else:
                    print(f"   ⚠️ Cảnh báo: Không tìm thấy sheet '{sheet_name}'")
        
        workbook.save(self.excel_output_path)
        # Lưu JSON để dùng cho bước tạo lời giải
        self.json_data_path.write_text(json.dumps(all_generated_data, ensure_ascii=False, indent=2), encoding='utf-8')
        return all_generated_data

    def run_explanation_generation(self, questions_data: Dict):
        """Bước 3: Tạo lời giải dựa trên các câu hỏi đã sinh ra"""
        print(f"\n--- [BƯỚC 3] TẠO LỜI GIẢI ---")
        exp_configs = get_explanation_config(self.hsk_level)
        if not exp_configs: return

        workbook = openpyxl.load_workbook(self.excel_output_path)
        
        # Làm phẳng dữ liệu JSON (Flatten) để tạo các task lời giải
        for prompt_id, data in questions_data.items():
            for q_type, q_content in data.items():
                if q_type not in exp_configs: continue
                
                config = exp_configs[q_type]
                print(f"   💡 Đang tạo lời giải cho dạng: {q_type}")
                
                # Làm phẳng dữ liệu và bọc trong 'data'
                tasks = []
                if isinstance(q_content, list):
                    for q in q_content:
                        tasks.append({"data": q})
                elif isinstance(q_content, dict):
                    if 'questions' in q_content:
                        shared_material = q_content.get('shared_material', [])
                        for q in q_content['questions']:
                            tasks.append({
                                "data": q,
                                "shared_material": shared_material
                            })
                    else:
                        tasks.append({"data": q_content})
                
                # Điểm mới: Tạo task list và gọi AI (có thể dùng ThreadPool ở đây để nhanh hơn)
                for task_item in tasks:
                    if 'prompt_file' not in config or 'schema_path' not in config:
                        # Bỏ qua các task không dùng API tạo lời giải (như các dạng của HSK1/HSK5 có sẵn đáp án)
                        continue

                    prompt_filename = config['prompt_file']
                    schema_filename = config['schema_path']
                    
                    # Xác định thư mục chứa file bằng cách lấy tiền tố từ tên file (vd: "hsk1_..." -> "hsk1")
                    prompt_folder = prompt_filename.split('_')[0]
                    schema_folder = schema_filename.split('_')[0]

                    # 1. Dùng Builder tạo Prompt text (không dùng file tạm)
                    prompt_file_text = self._load_resource(f"resources/prompts/explanation/{prompt_folder}/{prompt_filename}")
                    final_prompt = config['builder'](task_item, prompt_file_text)
                    
                    # 2. Load Schema
                    schema = self._load_resource(f"resources/schemas/explanation/{schema_folder}/{schema_filename}")
                    
                    # 3. Gọi AI lấy lời giải
                    explanation_result = ai_client.generate_content(final_prompt, schema)
                    
                    # 4. Dùng Renderer ghi vào Excel
                    sheet_name = config['sheet_name']
                    if sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        # Tìm hàng trống trong cột I
                        current_row = 2
                        while sheet[f'I{current_row}'].value is not None:
                            current_row += 1
                        
                        renderer = config['renderer']
                        cell = sheet[f'I{current_row}']
                        
                        # Xử lý các dạng khác nhau tùy theo số lượng tham số của renderer
                        import inspect
                        sig = inspect.signature(renderer)
                        if 'worksheet' in sig.parameters:
                            renderer(sheet, current_row, explanation_result, task_item)
                        elif len(sig.parameters) == 3:
                            renderer(cell, explanation_result, task_item)
                        else:
                            renderer(cell, explanation_result)

        # Xử lý đặc biệt cho sheet "ĐS (img) HSK1" và "ĐS Ko phụ đề (img) HSK1"
        if "ĐS (img) HSK1" in workbook.sheetnames:
            ds_worksheet = workbook["ĐS (img) HSK1"]
            render_ds_vocab_explanation(ds_worksheet, "ĐS (img) HSK1")
        if "ĐS Ko phụ đề (img) HSK1" in workbook.sheetnames:
            ds_no_sub_worksheet = workbook["ĐS Ko phụ đề (img) HSK1"]
            render_ds_vocab_explanation(ds_no_sub_worksheet, "ĐS Ko phụ đề (img) HSK1")

        workbook.save(self.excel_output_path)

    def start(self):
        """Khởi động toàn bộ Pipeline"""
        print(f"================ START PIPELINE: {self.hsk_level.upper()} ================")
        try:
            # B1: Load Matrix Data (Thay thế Preprocessing)
            print(f"\n--- [BƯỚC 1] LOAD MATRIX & VOCAB DATA ---")
            matrix_path = get_resource_path(f"resources/data/matrix_{self.hsk_level}.json")
            vocab_path = get_resource_path("resources/data/hsk_vocab.json")
            
            matrix_manager = MatrixManager(matrix_path=str(matrix_path), vocab_path=str(vocab_path))
            
            # B2: Questions
            questions_data = self.run_question_generation(matrix_manager)
            
            # B3: Explanations
            self.run_explanation_generation(questions_data)
            
            print(f"\n✨ TẤT CẢ HOÀN TẤT! ✨")
            print(f"📂 Kết quả: {self.excel_output_path}")
            
        except Exception as e:
            print(f"❌ PIPELINE BỊ LỖI: {e}")
            import traceback
            traceback.print_exc()

# --- TEST ---
from datetime import datetime
if __name__ == "__main__":
    # Test thử với HSK1
    pipeline = HSKPipeline(pdf_folder=r"input/tóm tắt HSK/hsk1", hsk_level="hsk1")
    pipeline.start()