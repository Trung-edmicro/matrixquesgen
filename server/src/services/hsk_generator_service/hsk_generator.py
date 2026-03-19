import json
import sys
import io
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Union, BinaryIO

try:
    import openpyxl
except ImportError:
    print("Thư viện 'openpyxl' chưa được cài. Hãy chạy: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Ánh xạ tên cột Excel -> key dữ liệu
# ---------------------------------------------------------------------------
# Chỉ số cột (0-based): A=0, B=1, C=2, E=4, F=5, G=6, I=8, K=10, M=12, O=14, P=15, Q=16, R=17, S=18
COLUMN_MAP = {
    0:  "ID",                 # Cột A  - Câu
    1:  "Phan",               # Cột B  - Phần
    2:  "Skill",              # Cột C  - Kỹ năng
    4:  "_dang_bai",          # Cột E  - Dạng bài  (dùng làm key nhóm, không xuất trực tiếp)
    5:  "So_cau_hoi",         # Cột F  - SL theo dạng bài
    6:  "Do_kho",             # Cột G  - Độ khó
    8:  "Trong_tam",          # Cột I  - Trọng tâm chính
    10: "Nhom_tu_vung",       # Cột K  - Nhóm từ vựng
    12: "Nhom_ngu_phap",      # Cột M  - Nhóm ngữ pháp
    14: "Chu-de_tinh-huong",  # Cột O  - Chủ đề / tình huống
    15: "Tu_trong_tam",       # Cột P  - Từ vựng trọng tâm
    16: "Ngu-phap_mau-cau",   # Cột Q  - Ngữ pháp / mẫu câu
    17: "Dap_an_mau",         # Cột R  - Đáp án mẫu
    18: "Dac_ta_tao_de_tt",   # Cột S  - Đặc tả tạo đề tương tự
}

# Thứ tự các key trong mỗi object data (không bao gồm _dang_bai)
DATA_KEYS_ORDER = [
    "ID", "Phan", "Skill", "So_cau_hoi", "Do_kho", "Trong_tam",
    "Nhom_tu_vung", "Nhom_ngu_phap", "Chu-de_tinh-huong",
    "Tu_trong_tam", "Ngu-phap_mau-cau", "Dap_an_mau", "Dac_ta_tao_de_tt",
]

# Hàng tiêu đề (1-based trong Excel = hàng 4)
HEADER_ROW = 4
# Dữ liệu bắt đầu từ hàng tiếp theo
DATA_START_ROW = HEADER_ROW + 1


def _cell_value(row_data: tuple, col_idx: int):
    """
    Lấy giá trị của ô tại vị trí col_idx (0-based) trong row_data.
    Trả về None nếu ô trống hoặc chỉ chứa khoảng trắng.
    """
    try:
        val = row_data[col_idx].value
    except IndexError:
        return None

    if val is None:
        return None
    if isinstance(val, str) and val.strip() == "":
        return None
        
    # Xử lý trường hợp số 1 bị Excel / Python hiểu nhầm thành số biến thực 1.0
    if isinstance(val, float) and val.is_integer():
        return int(val)
        
    # Trả về chuỗi đã strip nếu là str, nguyên giá trị nếu là số/ngày
    return val.strip() if isinstance(val, str) else val


def process_xlsx(file_input: Union[str, Path, BinaryIO], output_path: Optional[str] = None) -> dict:

    if isinstance(file_input, (str, Path)):
        path = Path(file_input)
        if not path.exists():
            raise FileNotFoundError(f"Không tìm thấy file: {path}")
        if path.suffix.lower() != ".xlsx":
            raise ValueError(f"File phải có đuôi .xlsx, nhận được: {path.suffix}")
        # Mở workbook, đọc sheet đầu tiên
        wb = openpyxl.load_workbook(path, data_only=True)
    else:
        # Giả định file_input là BytesIO / file-like object có đuôi xlsx
        wb = openpyxl.load_workbook(file_input, data_only=True)

    # 1. Tìm sheet có tên bắt đầu bằng "Ma trận_" để lấy Lever
    lever_value = "Unknown"
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("Ma trận_"):
            lever_value = sheet_name.split("_", 1)[1]
            target_sheet = wb[sheet_name]
            break
            
    # Nếu không tìm thấy sheet "Ma trận_", mặc định lấy sheet đầu tiên
    ws = target_sheet if target_sheet else wb.active

    # ----------------------------------------------------------------------
    # Đọc toàn bộ dữ liệu (từ hàng DATA_START_ROW trở đi) theo từng hàng
    # ----------------------------------------------------------------------

    # groups: list of {"key": str, "rows": [dict, ...]}
    groups: List[dict] = []
    current_key: Optional[str] = None  # noqa: F841

    for row in ws.iter_rows(min_row=DATA_START_ROW):
        dang_bai = _cell_value(row, 4)  # Cột E (index 4)

        # Bỏ qua hàng không có dạng bài
        if dang_bai is None:
            continue

        # Xây dựng object dữ liệu cho hàng này
        record: dict = {}
        for col_idx, key in COLUMN_MAP.items():
            if key == "_dang_bai":
                continue  # không thêm vào record
            record[key] = _cell_value(row, col_idx)

        # Gom theo dạng bài LIÊN TIẾP
        if groups and groups[-1]["key"] == dang_bai:
            groups[-1]["rows"].append(record)
        else:
            groups.append({"key": dang_bai, "rows": [record]})
            current_key = dang_bai

    # ----------------------------------------------------------------------
    # Xây dựng JSON output
    # ----------------------------------------------------------------------
    result: dict = {}

    for group in groups:
        key = group["key"]
        rows = group["rows"]

        # Thứ tự key trong record theo DATA_KEYS_ORDER
        ordered_rows = []
        for r in rows:
            ordered_record = {k: r.get(k) for k in DATA_KEYS_ORDER}
            ordered_rows.append(ordered_record)

        # Nếu key đã tồn tại (cùng tên nhưng không liên tiếp),
        # nối thêm vào data và cập nhật SL_dang_bai
        if key in result:
            result[key]["data"].extend(ordered_rows)
            result[key]["SL_dang_bai"] = len(result[key]["data"])
        else:
            result[key] = {
                "SL_dang_bai": len(ordered_rows),
                "data": ordered_rows,
            }

    # Đưa key Lever lên cùng của object JSON trả về
    final_result = {"Lever": lever_value}
    final_result.update(result)

    # ----------------------------------------------------------------------
    # Ghi file JSON nếu có output_path
    # ----------------------------------------------------------------------
    if output_path:
        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        with open(out, "w", encoding="utf-8") as f:
            json.dump(final_result, f, ensure_ascii=False, indent=2)
        print(f"✅ Đã ghi kết quả ra: {out.resolve()}")

    return final_result

from fastapi import HTTPException
import logging

logger = logging.getLogger(__name__)

async def generate_hsk_flow(file):
    """
    Xử lý logic sinh câu hỏi cho HSK (Tiếng Trung)
    """
    logger.info(f"Bắt đầu xử lý file HSK: {file.filename}")
    
    try:
        content = await file.read()
        file_obj = io.BytesIO(content)
        data = process_xlsx(file_obj)
        
        # --- LOG KẾT QUẢ PARSE DỮ LIỆU ---
        logger.info("=== KẾT QUẢ PARSE DỮ LIỆU ===")
        # In toàn bộ data gốc theo format JSON dễ nhìn
        print(json.dumps(data, ensure_ascii=False, indent=2))
        logger.info("=============================")
        # --------------------------------

        # Tạm thời trả về kết quả parse để kiểm tra
        return {
            "session_id": "test-session",
            "status": "success",
            "message": "Đã parse file HSK thành công.",
            "data": data  # Trả về toàn bộ data kèm các câu hỏi chi tiết
        }
        
    except Exception as e:
        logger.error(f"Lỗi xử lý HSK: {e}")
        raise HTTPException(
            status_code=500, 
            detail=f"Lỗi xử lý file HSK: {str(e)}"
        )
